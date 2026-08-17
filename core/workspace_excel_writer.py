import datetime
import os
from typing import List, Dict, Any, Union

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from models.bom_item import BomItem
from models.workspace import Workspace
from services.project_aggregation import WorkspaceAggregationResult, WorkspaceAggregatedComponent
from core.base_excel_writer import BaseExcelWriter


class WorkspaceExcelWriter(BaseExcelWriter):
    """Writes a comprehensive workspace-aware multi-project BOM cost report to Excel."""

    def __init__(
        self,
        workspace: Workspace,
        aggregation_result: WorkspaceAggregationResult,
        enriched_items: List[BomItem],
        component_keys: List[str],
        build_multipliers: List[int] = None,
        pricing_mode: str = "unit",
    ):
        BaseExcelWriter.__init__(self, pricing_mode=pricing_mode)
        
        # 1. Validation: Length match
        if len(enriched_items) != len(component_keys):
            raise ValueError(f"Length mismatch: enriched_items({len(enriched_items)}) != component_keys({len(component_keys)})")

        # 2. Validation: Duplicates
        if len(set(component_keys)) != len(component_keys):
            raise ValueError("Duplicate component keys found in component_keys list.")

        # 3. Validation: Exact match
        agg_keys = {c.component_key for c in aggregation_result.components}
        prov_keys = set(component_keys)
        if agg_keys != prov_keys:
            missing = agg_keys - prov_keys
            extra = prov_keys - agg_keys
            raise ValueError(f"Component key mismatch. Missing: {missing}, Extra: {extra}")

        self.workspace = workspace
        self.aggregation_result = aggregation_result
        self.enriched_items = enriched_items
        self.component_keys = component_keys
        self.build_multipliers = build_multipliers or [1, 5, 10, 50, 100]

        # 4. Dictionary Mapping
        self.enriched_by_key: Dict[str, BomItem] = dict(zip(component_keys, enriched_items))
        
        self.component_by_key: Dict[str, WorkspaceAggregatedComponent] = {
            c.component_key: c for c in self.aggregation_result.components
        }
        
        self.item_id_to_key = {
            id(usage.item): comp.component_key
            for comp in self.aggregation_result.components
            for usage in comp.usages
        }

    def _get_pricing(self, comp_key: str, multiplied_qty: Union[int, float]) -> Dict[str, Any]:
        item = self.enriched_by_key[comp_key]
        return self._get_pricing_for_component_item(item, multiplied_qty)

    def write(self, output_path: str):
        used_names = set()

        # Keep the workbook's primary overview sheets first. Excel opens the
        # first created worksheet by default, and the explicit active-sheet
        # assignment below protects that order if more sheets are added later.
        self._add_aggregated_components_sheet(used_names)
        self._add_project_sheets(used_names)

        # A single source BOM can contain multiple Board/Kart values. Each
        # actual board still receives its own worksheet and supplier prices.
        self._add_board_sheets(used_names)
        self._add_mutual_components_sheet(used_names)
        self._add_supplier_stock_sheet(self.enriched_items)

        self.wb.active = self.wb.sheetnames.index("All Aggregated")
        
        self.wb.save(output_path)

    def _add_board_sheets(self, used_names: set[str]):
        """Write each actual Board/Kart value to its own worksheet."""
        headers = [
            "Board Name", "Description", "Designator", "Quantity (Per Board / Total)",
            "Value", "Manufacturer", "MPN", "Design Item ID", "JLCPCB Part Number",
            "DigiKey Part Number", "JLCPCB Stock", "DigiKey Stock", "JLCPCB Unit Price",
            "DigiKey Unit Price", "Board Required Quantity", "Pricing Pool Quantity",
            "JLCPCB Total Price", "DigiKey Total Price", "Selected Supplier",
            "Mixed Sourcing Total Price", "Status",
        ]
        build_quantity = self.build_multipliers[0]

        board_groups = {}
        for comp_key, comp in self.component_by_key.items():
            for usage in comp.usages:
                key = (usage.project_name, usage.board_name)
                board_groups.setdefault(key, {}).setdefault(comp_key, []).append(usage)

        for (project_name, board_name), component_usages in board_groups.items():
            sheet_name = self._safe_sheet_name(board_name, used_names)
            ws = self.wb.create_sheet(sheet_name)
            all_board_usages = [
                usage
                for usages in component_usages.values()
                for usage in usages
            ]
            source_files = sorted({
                os.path.basename(usage.board_file_path)
                for usage in all_board_usages
            })
            board_instances_per_set = sum({
                usage.board_file_path: usage.board_quantity
                for usage in all_board_usages
            }.values())
            production_quantity = build_quantity * board_instances_per_set

            ws.append(["Project:", project_name])
            ws.append(["Board Name:", board_name])
            ws.append(["Source BOM File(s):", ", ".join(source_files)])
            ws.append(["Production Quantity:", production_quantity])
            ws.append(["Price Tier Pool:", "All boards in this project using the same MPN"])
            ws.append([])
            ws.append(headers)
            header_row = ws.max_row
            data_start_row = header_row + 1
            for cell in ws[header_row]:
                cell.font = self.header_font
                cell.alignment = self.header_alignment

            missing_price_count = 0
            for comp_key, board_usages in component_usages.items():
                comp = self.component_by_key[comp_key]
                enriched = self.enriched_by_key[comp_key]
                board_required_quantity = sum(u.total_quantity for u in board_usages) * build_quantity
                pricing_pool_quantity = sum(
                    u.total_quantity for u in comp.usages if u.project_name == project_name
                ) * build_quantity
                _, j_price, d_price, _, _ = self._component_price_values(
                    enriched, pricing_pool_quantity
                )
                j_total = board_required_quantity * j_price if j_price is not None else None
                d_total = board_required_quantity * d_price if d_price is not None else None
                selected_source, selected_price = self._selected_supplier_price(
                    enriched, j_price, d_price
                )
                mixed_total = (
                    board_required_quantity * selected_price
                    if selected_price is not None else None
                )
                if selected_price is None:
                    missing_price_count += 1

                for usage_index, usage in enumerate(board_usages):
                    ws.append([
                        board_name,
                        enriched.description,
                        usage.item.designator,
                        self._fmt_qty_pair(usage.bom_line_quantity, usage.total_quantity),
                        enriched.value,
                        enriched.manufacturer,
                        enriched.mpn,
                        usage.item.comment,
                        enriched.jlcpcb_part_number,
                        enriched.digikey_part_number,
                        enriched.available_stock_qty if enriched.available_stock_qty is not None else "-",
                        enriched.digikey_stock_qty if enriched.digikey_stock_qty is not None else "-",
                        j_price,
                        d_price,
                        board_required_quantity,
                        pricing_pool_quantity,
                        j_total if usage_index == 0 else None,
                        d_total if usage_index == 0 else None,
                        selected_source if usage_index == 0 else None,
                        mixed_total if usage_index == 0 else None,
                        enriched.status,
                    ])
                    self._format_supplier_price_cells(ws, ws.max_row, 13, 14, enriched)
                    for total_col in (17, 18):
                        if isinstance(ws.cell(row=ws.max_row, column=total_col).value, (int, float)):
                            self._format_currency(ws.cell(row=ws.max_row, column=total_col))
                    if isinstance(ws.cell(row=ws.max_row, column=20).value, (int, float)):
                        self._format_currency(ws.cell(row=ws.max_row, column=20))
                    status_cell = ws.cell(row=ws.max_row, column=21)
                    fill = self._get_status_fill(enriched.status, bool(enriched.jlcpcb_part_number))
                    if fill:
                        status_cell.fill = fill

            data_end_row = ws.max_row
            self._add_price_totals_box(
                ws,
                header_row,
                data_start_row,
                data_end_row,
                build_quantity,
                total_label="Board Total",
                missing_price_count=missing_price_count,
            )
            self._add_board_cost_summary(
                ws, header_row, data_start_row, data_end_row, production_quantity,
                missing_price_count=missing_price_count,
            )
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{data_end_row}"
            ws.freeze_panes = f"A{data_start_row}"
            self._auto_fit_columns(ws)
            self._apply_narrow_columns(ws, header_row=header_row)

    def _add_aggregated_components_sheet(self, used_names: set[str]):
        sheet_name = self._safe_sheet_name("All Aggregated", used_names)
        ws = self.wb.create_sheet(sheet_name)
        
        headers = [
            "Board Name", "Description", "Designator", "Quantity (Per Board / Total)", 
            "Value", "Manufacturer", "MPN", "Design Item ID", "JLCPCB Part Number", "DigiKey Part Number", "JLCPCB Stock", "DigiKey Stock",
            "JLCPCB Unit Price", "DigiKey Unit Price", "Pricing Quantity", "JLCPCB Total Price", "DigiKey Total Price", "Status"
        ]

        ws.append(headers)
        data_start_row = 2
        for cell in ws[1]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            
        for comp_key, comp in self.component_by_key.items():
            enriched = self.enriched_by_key[comp_key]
            pricing_quantity, j_price, d_price, j_total, d_total = self._component_price_values(enriched)
            
            for usage_index, u in enumerate(comp.usages):
                board_str = u.board_name
                desigs_str = str(u.item.designator) if u.item.designator else ""
                qty_str = self._fmt_qty_pair(u.bom_line_quantity, u.total_quantity)
                
                row = [
                    board_str,
                    enriched.description,
                    desigs_str,
                    qty_str,
                    enriched.value,
                    enriched.manufacturer,
                    enriched.mpn,
                    enriched.comment,
                    enriched.jlcpcb_part_number,
                    enriched.digikey_part_number,
                    enriched.available_stock_qty if enriched.available_stock_qty is not None else "-",
                    enriched.digikey_stock_qty if enriched.digikey_stock_qty is not None else "-",
                    j_price,
                    d_price,
                    pricing_quantity,
                    j_total if usage_index == 0 else None,
                    d_total if usage_index == 0 else None,
                    enriched.status
                ]
                    
                ws.append(row)
                ws.cell(row=ws.max_row, column=1).alignment = self.wrap_alignment
                ws.cell(row=ws.max_row, column=4).alignment = self.wrap_alignment
                
                self._format_supplier_price_cells(ws, ws.max_row, 13, 14, enriched)
                for total_col in (16, 17):
                    if isinstance(ws.cell(row=ws.max_row, column=total_col).value, (int, float)):
                        self._format_currency(ws.cell(row=ws.max_row, column=total_col))
                cell = ws.cell(row=ws.max_row, column=18)
                fill = self._get_status_fill(enriched.status, bool(enriched.jlcpcb_part_number))
                if fill:
                    cell.fill = fill

        data_end_row = ws.max_row
        missing_price_count = sum(
            1
            for enriched in self.enriched_by_key.values()
            if self._selected_supplier_price(
                enriched,
                self._component_price_values(enriched)[1],
                self._component_price_values(enriched)[2],
            )[1] is None
        )
        self._add_price_totals_box(
            ws, 1, data_start_row, data_end_row, self.build_multipliers[0],
            total_label="Workspace Total",
            missing_price_count=missing_price_count,
        )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end_row}"
        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws)
        self._apply_narrow_columns(ws)

    def _add_mutual_components_sheet(self, used_names: set[str]):
        sheet_name = self._safe_sheet_name("Mutual Components", used_names)
        ws = self.wb.create_sheet(sheet_name)
        
        headers = [
            "Board Names", "Description", "Designator", "Quantity (Per Board / Total)", 
            "Value", "Manufacturer", "MPN", "Design Item ID", "JLCPCB Part Number", "DigiKey Part Number", "JLCPCB Stock", "DigiKey Stock",
            "JLCPCB Unit Price", "DigiKey Unit Price", "Pricing Quantity", "JLCPCB Total Price", "DigiKey Total Price", "Status"
        ]

        ws.append(headers)
        data_start_row = 2
        for cell in ws[1]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            
        for mutual in self.aggregation_result.mutual_components:
            comp_key = mutual.component_key
            enriched = self.enriched_by_key[comp_key]
            pricing_quantity, j_price, d_price, j_total, d_total = self._component_price_values(enriched)
            
            board_str = "\n".join([u.board_name for u in mutual.usages])
            qty_str = self._format_usage_quantities(mutual.usages)
            
            row = [
                board_str,
                enriched.description,
                "Mutual", # Designators can be huge for mutual, so we just say Mutual
                qty_str,
                enriched.value,
                enriched.manufacturer,
                enriched.mpn,
                enriched.comment,
                enriched.jlcpcb_part_number,
                enriched.digikey_part_number,
                enriched.available_stock_qty if enriched.available_stock_qty is not None else "-",
                enriched.digikey_stock_qty if enriched.digikey_stock_qty is not None else "-",
                j_price,
                d_price,
                pricing_quantity,
                j_total,
                d_total,
                enriched.status
            ]
                
            ws.append(row)
            ws.cell(row=ws.max_row, column=1).alignment = self.wrap_alignment
            ws.cell(row=ws.max_row, column=4).alignment = self.wrap_alignment
            
            self._format_supplier_price_cells(ws, ws.max_row, 13, 14, enriched)
            for total_col in (16, 17):
                if isinstance(ws.cell(row=ws.max_row, column=total_col).value, (int, float)):
                    self._format_currency(ws.cell(row=ws.max_row, column=total_col))
            cell = ws.cell(row=ws.max_row, column=18)
            fill = self._get_status_fill(enriched.status, bool(enriched.jlcpcb_part_number))
            if fill:
                cell.fill = fill

        data_end_row = ws.max_row
        missing_price_count = sum(
            1
            for mutual in self.aggregation_result.mutual_components
            if self._selected_supplier_price(
                self.enriched_by_key[mutual.component_key],
                self._component_price_values(
                    self.enriched_by_key[mutual.component_key]
                )[1],
                self._component_price_values(
                    self.enriched_by_key[mutual.component_key]
                )[2],
            )[1] is None
        )
        self._add_price_totals_box(
            ws, 1, data_start_row, data_end_row, self.build_multipliers[0],
            total_label="Mutual Components Total",
            missing_price_count=missing_price_count,
        )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end_row}"
        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws)
        self._apply_narrow_columns(ws)

    def _add_project_sheets(self, used_names: set[str]):
        for project_name, proj_result in self.aggregation_result.project_results.items():
            base_name = f"Project Summary - {project_name}"
            sheet_name = self._safe_sheet_name(base_name, used_names)
            ws = self.wb.create_sheet(sheet_name)
            
            project_obj = self.workspace.get_project(project_name)
            actual_board_names = {
                usage.board_name
                for component in proj_result.components
                for usage in component.usages
            }
            
            # Count mutual components appearing in this project
            mutual_count = sum(1 for m in self.aggregation_result.mutual_components if project_name in self.component_by_key[m.component_key].source_projects)
            
            ws.append(["Project Name:", project_name])
            ws.append(["Number of Boards:", len(actual_board_names)])
            ws.append(["Number of BOM Files:", len(project_obj.board_items) if project_obj else 0])
            ws.append(["Unique Components in Project:", len(proj_result.components)])
            ws.append(["Mutual Workspace Components in Project:", mutual_count])
            ws.append([
                "Mixed Sourcing Logic:",
                "Use JLCPCB when its code and price exist; use DigiKey only as fallback. "
                "The same component is never counted twice.",
            ])
            ws.append([])
            
            for row_idx in range(1, 7):
                ws.cell(row=row_idx, column=1).font = self.header_font

            # Cost Summary
            ws.append(["Project Cost Summary"])
            ws[f"A{ws.max_row}"].font = Font(bold=True, size=14)
            ws.append([
                "Build Qty", "JLCPCB Cost", "DigiKey Fallback Cost",
                "Mixed Sourcing Total", "DigiKey-Only Total",
                "Missing Price Count", "Cost Status",
            ])
            for cell in ws[ws.max_row]:
                cell.font = self.header_font
                cell.alignment = self.header_alignment

            for m in self.build_multipliers:
                tot_jlc = 0.0
                tot_rem_dk = 0.0
                tot_comb = 0.0
                tot_dk_only = 0.0
                missing_price_count = 0
                
                for proj_comp in proj_result.components:
                    comp_key = proj_comp.component_key
                    # We price based on the project's specific localized sub-quantity to get accurate breakpoints
                    mult_qty = proj_comp.total_quantity * m
                    pricing = self._get_pricing(comp_key, mult_qty)
                    
                    if pricing["jlcpcb_cost"] is not None: tot_jlc += pricing["jlcpcb_cost"]
                    if pricing["remaining_digikey_cost"] is not None: tot_rem_dk += pricing["remaining_digikey_cost"]
                    if pricing["combined_cost"] is not None: tot_comb += pricing["combined_cost"]
                    else: missing_price_count += 1
                    if pricing["digikey_only_cost"] is not None: tot_dk_only += pricing["digikey_only_cost"]

                cost_status = "COMPLETE" if missing_price_count == 0 else "INCOMPLETE"
                row = [
                    f"{m}x", tot_jlc, tot_rem_dk, tot_comb, tot_dk_only,
                    missing_price_count, cost_status,
                ]
                ws.append(row)
                for c_idx in range(2, 6):
                    self._format_currency(ws.cell(row=ws.max_row, column=c_idx))
                for c_idx in (6, 7):
                    ws.cell(row=ws.max_row, column=c_idx).fill = (
                        self.fill_green if missing_price_count == 0 else self.fill_red
                    )

            ws.append([])
            
            # Component Table
            # Component Table
            headers = [
                "Board Name", "Description", "Designator", "Quantity (Per Board / Total)", 
                "Value", "Manufacturer", "MPN", "Design Item ID", "JLCPCB Part Number", "DigiKey Part Number", "JLCPCB Stock", "DigiKey Stock",
                "JLCPCB Unit Price", "DigiKey Unit Price", "Pricing Quantity", "JLCPCB Total Price", "DigiKey Total Price", "Status"
            ]
                
            ws.append(headers)
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                
            for proj_comp in proj_result.components:
                comp_key = proj_comp.component_key
                enriched = self.enriched_by_key[comp_key]
                display_quantity = proj_comp.total_quantity * self.build_multipliers[0]
                pricing_quantity, j_price, d_price, j_total, d_total = self._component_price_values(enriched, display_quantity)
                
                for usage_index, u in enumerate(proj_comp.usages):
                    board_str = u.board_name
                    desigs_str = str(u.bom_item.designator) if u.bom_item.designator else ""
                    qty_str = self._fmt_qty_pair(u.bom_line_quantity, u.total_quantity)
                    
                    row = [
                        board_str,
                        enriched.description,
                        desigs_str,
                        qty_str,
                        enriched.value,
                        enriched.manufacturer,
                        enriched.mpn,
                        enriched.comment,
                        enriched.jlcpcb_part_number,
                        enriched.digikey_part_number,
                        enriched.available_stock_qty if enriched.available_stock_qty is not None else "-",
                        enriched.digikey_stock_qty if enriched.digikey_stock_qty is not None else "-",
                        j_price,
                        d_price,
                        pricing_quantity,
                        j_total if usage_index == 0 else None,
                        d_total if usage_index == 0 else None,
                        enriched.status
                    ]
                        
                    ws.append(row)
                    ws.cell(row=ws.max_row, column=1).alignment = self.wrap_alignment
                    ws.cell(row=ws.max_row, column=4).alignment = self.wrap_alignment
                    
                    self._format_supplier_price_cells(ws, ws.max_row, 13, 14, enriched)
                    for total_col in (16, 17):
                        if isinstance(ws.cell(row=ws.max_row, column=total_col).value, (int, float)):
                            self._format_currency(ws.cell(row=ws.max_row, column=total_col))
                    cell = ws.cell(row=ws.max_row, column=18)
                    fill = self._get_status_fill(enriched.status, bool(enriched.jlcpcb_part_number))
                    if fill:
                        cell.fill = fill
                        
            data_start_row = header_row + 1
            data_end_row = ws.max_row
            missing_price_count = sum(
                1
                for proj_comp in proj_result.components
                if self._get_pricing(
                    proj_comp.component_key,
                    proj_comp.total_quantity * self.build_multipliers[0],
                )["combined_cost"] is None
            )
            self._add_price_totals_box(
                ws, header_row, data_start_row, data_end_row,
                self.build_multipliers[0],
                missing_price_count=missing_price_count,
            )
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{data_end_row}"
            ws.freeze_panes = f"A{header_row + 1}"
            self._auto_fit_columns(ws)
            self._apply_narrow_columns(ws, header_row=header_row)

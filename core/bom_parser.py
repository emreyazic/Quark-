"""BOM file parser with auto-detection of column mappings.

Fixes vs. original:
- MPN patterns are checked BEFORE manufacturer to prevent 'Manufacturer Part Number'
  from being mapped to the manufacturer field.
- Manufacturer regex uses a negative lookahead to exclude 'Manufacturer Part Number'.
- Content-based heuristic detects when manufacturer and MPN columns appear swapped.
- Duplicate column mapping (same column to two fields) is detected and warned.
"""

import os
import re
from typing import Optional

import openpyxl

from models.bom_item import BomFile, BomItem, ColumnMapping
from core.mpn_utils import (
    clean_mpn_value,
    is_mpn_like,
    is_manufacturer_like,
    parse_positive_integer_quantity,
)


# Known header patterns for auto-detection.
# IMPORTANT: Order matters — more specific patterns (MPN) must be checked
# before less specific ones (Manufacturer) to avoid ambiguity.
# Each tuple is (field_name, patterns_list, priority) where lower priority = checked first.
HEADER_PATTERNS_ORDERED = [
    # MPN patterns — checked FIRST (priority 0)
    ("mpn", [
        r"manufacturer\s*part\s*number",
        r"mfr\s*part\s*(?:number|no|num|#)",
        r"\bmpn\b",
        r"(?<!\bmanufacturer\b\s)part\s*number",
        r"mfg\s*part\s*(?:number|no|num|#)",
        r"mfr\s*p/n",
        r"man\.?\s*part\s*(?:number|no|num|#)?",
    ]),
    # Quantity
    ("quantity", [
        r"\bquantity\b",
        r"\bqty\b",
        r"\bcount\b",
        r"\badet\b",
        r"\bmiktar\b",
    ]),
    # Manufacturer — with negative lookahead to avoid matching "Manufacturer Part Number"
    ("manufacturer", [
        r"\bmanufacturer\b(?!\s*part)",
        r"\bmfr\b(?!\s*part)",
        r"\bmfg\b(?!\s*part)",
        r"\büretici\b",
    ]),
    # Description
    ("description", [
        r"\bdescription\b",
        r"\bdesc\b",
        r"\baciklama\b",
        r"\baçıklama\b",
    ]),
    # Designator
    ("designator", [
        r"\bdesignator\b",
        r"\bref\s*des\b",
        r"\breference\b",
    ]),
    # Comment
    ("comment", [
        r"\bcomment\b",
        r"\bcomponent\b",
        r"\bnot\b",
        r"\byorum\b",
        r"\bname\b",
    ]),
    # Footprint
    ("footprint", [
        r"\bfootprint\b",
        r"\bpackage\b",
        r"\bkılıf\b",
    ]),
    # Value
    ("value", [
        r"\bvalue\b",
        r"\bdeger\b",
        r"\bdeğer\b",
    ]),
    # Board Identifier (e.g. Kart)
    ("board_identifier", [
        r"\bkart\b",
        r"\bboard\b",
        r"\bpcb\b",
    ]),
]


class BomParser:
    """Parses Altium Designer BOM Excel files."""

    def __init__(self):
        pass

    @staticmethod
    def _validate_file_extension(file_path: str) -> None:
        extension = os.path.splitext(file_path)[1].lower()
        if extension == ".xls":
            raise ValueError(
                "Legacy .xls files are not supported. Save the workbook as "
                ".xlsx and try again."
            )
        if extension not in {".xlsx", ".xlsm"}:
            raise ValueError(
                f"Unsupported BOM file type '{extension or '(none)'}'. "
                "Use an .xlsx or .xlsm workbook."
            )

    @staticmethod
    def _read_sheet_preview(ws) -> tuple[list[str], list[list[str]], int]:
        """Return headers, a five-row preview, and non-empty data-row count."""
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[1]
        ]
        while headers and headers[-1] == "":
            headers.pop()

        preview_rows = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = list(row[: len(headers)])
            if all(v is None or str(v).strip() == "" for v in row_data):
                continue
            row_count += 1
            if len(preview_rows) < 5:
                preview_rows.append(
                    [str(v) if v is not None else "" for v in row_data]
                )
        return headers, preview_rows, row_count

    def inspect_sheets(self, file_path: str) -> list[BomFile]:
        """Inspect all visible sheets in a workbook and return a BomFile for each.

        Detects headers, row counts, preview rows, auto-detected column mappings,
        and flags probable duplicate sheets.

        Args:
            file_path: Path to the Excel file.

        Returns:
            List of BomFile objects, one for each visible worksheet.
        """
        self._validate_file_extension(file_path)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        inspected_sheets: list[BomFile] = []
        sheet_signatures: list[tuple[str, int, tuple, tuple]] = []

        try:
            for ws in wb.worksheets:
                if ws.sheet_state != "visible":
                    continue
                headers, preview_rows, row_count = self._read_sheet_preview(ws)
                mapping = self._auto_detect_columns(headers, preview_rows)

                # Duplicate sheet detection
                # Signature: (headers, row_count, tuple of preview rows)
                preview_sig = tuple(tuple(r) for r in preview_rows)
                headers_sig = tuple(headers)
                duplicate_of = None

                # Check if identical to an already inspected sheet
                if row_count > 0 or headers:
                    for prev_title, prev_rc, prev_hdrs, prev_prev in sheet_signatures:
                        # Exact duplicate: matching row count, headers, and preview rows
                        if (
                            row_count == prev_rc
                            and headers_sig == prev_hdrs
                            and preview_sig == prev_prev
                        ):
                            duplicate_of = prev_title
                            break
                        # High similarity: same title prefix (e.g. "Sheet1" vs "Sheet1 (Copy)")
                        # and matching row count
                        title_clean = re.sub(
                            r"[\s_]*\((?:copy|\d+)\)|[\s_]+copy\b|[\s_]+kopya\b",
                            "",
                            ws.title,
                            flags=re.IGNORECASE,
                        ).strip()
                        prev_clean = re.sub(
                            r"[\s_]*\((?:copy|\d+)\)|[\s_]+copy\b|[\s_]+kopya\b",
                            "",
                            prev_title,
                            flags=re.IGNORECASE,
                        ).strip()
                        if (
                            title_clean.lower() == prev_clean.lower()
                            and row_count == prev_rc
                            and row_count > 0
                        ):
                            duplicate_of = prev_title
                            break

                sheet_signatures.append(
                    (ws.title, row_count, headers_sig, preview_sig)
                )

                warnings = list(mapping.warnings)
                if duplicate_of:
                    warn_msg = (
                        f"⚠ Probable duplicate of sheet '{duplicate_of}' "
                        f"(identical or near-identical BOM content detected)."
                    )
                    warnings.append(warn_msg)
                    mapping.warnings.append(warn_msg)

                bom_file = BomFile(
                    file_path=file_path,
                    board_name=base_name,
                    sheet_name=ws.title,
                    headers=headers,
                    column_mapping=mapping,
                    row_count=row_count,
                    preview_rows=preview_rows,
                    is_valid=mapping.is_valid(),
                    duplicate_of=duplicate_of,
                    warnings=warnings,
                )
                inspected_sheets.append(bom_file)

            if not inspected_sheets:
                raise ValueError("The workbook contains no visible worksheets.")

            return inspected_sheets
        finally:
            wb.close()

    def load_file(self, file_path: str, sheet_name: Optional[str] = None) -> BomFile:
        """Load an Excel BOM file and return a BomFile with auto-detected mapping.

        Args:
            file_path: Path to the Excel file.
            sheet_name: Optional specific worksheet name to load. If None,
                selects the active/primary valid sheet.

        Returns:
            BomFile with headers, preview rows, and auto-detected column mapping.
        """
        sheets = self.inspect_sheets(file_path)

        if sheet_name is not None:
            for s in sheets:
                if s.sheet_name == sheet_name:
                    return s
            raise ValueError(
                f"Worksheet '{sheet_name}' not found in '{file_path}'."
            )

        # Look for active sheet first if valid and non-duplicate
        self._validate_file_extension(file_path)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            active_title = wb.active.title if wb.active else None
        finally:
            wb.close()

        if active_title:
            active_sheet = next(
                (s for s in sheets if s.sheet_name == active_title and s.is_valid and not s.duplicate_of),
                None,
            )
            if active_sheet:
                return active_sheet

        # Prefer first valid non-duplicate sheet
        primary = next((s for s in sheets if s.is_valid and not s.duplicate_of), None)
        if primary is None:
            primary = next((s for s in sheets if s.is_valid), sheets[0])

        return primary

    def _auto_detect_columns(
        self, headers: list[str], preview_rows: list[list]
    ) -> ColumnMapping:
        """Auto-detect column mapping from header names.

        Uses ordered pattern matching (MPN first) and negative lookaheads
        to prevent 'Manufacturer Part Number' from being mapped to manufacturer.

        Args:
            headers: List of header strings from the first row.
            preview_rows: First few data rows for content-based validation.

        Returns:
            ColumnMapping with detected column indices, confidence, and warnings.
        """
        mapping = ColumnMapping()
        matches_found = 0
        total_fields = len(HEADER_PATTERNS_ORDERED)
        used_columns: dict[int, str] = {}  # col_idx -> field_name that claimed it

        for field_name, patterns in HEADER_PATTERNS_ORDERED:
            best_col = None
            for col_idx, header in enumerate(headers):
                header_lower = header.lower().strip()
                if not header_lower:
                    continue
                # Skip columns already claimed by a higher-priority field
                if col_idx in used_columns:
                    continue
                for pattern in patterns:
                    if re.search(pattern, header_lower, re.IGNORECASE):
                        best_col = col_idx
                        break
                if best_col is not None:
                    break

            if best_col is not None:
                setattr(mapping, field_name, best_col)
                used_columns[best_col] = field_name
                matches_found += 1

        mapping.confidence = matches_found / total_fields if total_fields > 0 else 0.0

        # Content-based validation: detect manufacturer/MPN column swap
        self._validate_mfr_mpn_mapping(mapping, preview_rows, headers)

        return mapping

    def _validate_mfr_mpn_mapping(
        self,
        mapping: ColumnMapping,
        preview_rows: list[list],
        headers: list[str],
    ) -> None:
        """Check if manufacturer and MPN columns appear swapped using heuristics.

        If the 'manufacturer' column mostly contains MPN-like values and the
        'mpn' column mostly contains manufacturer-like values, add a warning.
        """
        if mapping.mpn is None or mapping.manufacturer is None:
            return
        if not preview_rows:
            return

        mpn_col = mapping.mpn
        mfr_col = mapping.manufacturer

        # Sample values from each column
        mpn_values = []
        mfr_values = []
        for row in preview_rows:
            if mpn_col < len(row):
                mpn_values.append(str(row[mpn_col]).strip())
            if mfr_col < len(row):
                mfr_values.append(str(row[mfr_col]).strip())

        # Count how many values in each column look like MPNs vs manufacturer names
        mpn_col_is_mpn = sum(1 for v in mpn_values if v and is_mpn_like(v))
        mpn_col_is_mfr = sum(1 for v in mpn_values if v and is_manufacturer_like(v))
        mfr_col_is_mpn = sum(1 for v in mfr_values if v and is_mpn_like(v))
        mfr_col_is_mfr = sum(1 for v in mfr_values if v and is_manufacturer_like(v))

        total_samples = len(preview_rows)
        if total_samples == 0:
            return

        # Detect likely swap: MPN column has manufacturer names AND manufacturer
        # column has part numbers
        swap_detected = (
            mpn_col_is_mfr > mpn_col_is_mpn
            and mfr_col_is_mpn > mfr_col_is_mfr
        )

        if swap_detected:
            mapping.warnings.append(
                f"⚠ Possible column swap detected: '{headers[mpn_col]}' (mapped as MPN) "
                f"contains manufacturer-like values, and '{headers[mfr_col]}' (mapped as "
                f"Manufacturer) contains MPN-like values. Please verify the mapping."
            )
            # Auto-swap to correct the mapping
            mapping.mpn, mapping.manufacturer = mapping.manufacturer, mapping.mpn
            mapping.warnings.append(
                "↔ Auto-swapped MPN and Manufacturer columns based on content analysis."
            )

    def parse_bom_items(self, bom_file: BomFile) -> list[BomItem]:
        """Parse rows from a single targeted BOM worksheet into BomItem objects.

        Only parses the worksheet specified by bom_file.sheet_name. Does NOT
        silently combine rows from other worksheets in the workbook.

        Args:
            bom_file: A BomFile with a valid column mapping and sheet_name.

        Returns:
            List of BomItem objects for that worksheet.

        Raises:
            ValueError: If column mapping is missing, invalid, or contains duplicate
                mapped columns, or if required fields are missing.
        """
        if bom_file.column_mapping is None:
            raise ValueError(
                f"Column mapping is missing for {bom_file.file_path}."
            )

        if bom_file.column_mapping.has_duplicate_mappings():
            dup_info = bom_file.column_mapping.get_duplicate_fields()
            dup_desc = ", ".join(
                f"Col {col+1} mapped to {fields}" for col, fields in dup_info.items()
            )
            raise ValueError(
                f"Invalid column mapping for {bom_file.file_path}: duplicate column mapping detected ({dup_desc}). "
                "Each spreadsheet column must be mapped to at most one field."
            )

        if not bom_file.column_mapping.is_valid():
            raise ValueError(
                f"Column mapping is incomplete for {bom_file.file_path}. "
                f"MPN and Quantity columns are required."
            )

        mapping = bom_file.column_mapping
        items = []

        self._validate_file_extension(bom_file.file_path)
        wb = openpyxl.load_workbook(bom_file.file_path, read_only=True, data_only=True)
        source_file_name = os.path.basename(bom_file.file_path)
        try:
            ws = None
            if bom_file.sheet_name and bom_file.sheet_name in wb.sheetnames:
                candidate = wb[bom_file.sheet_name]
                if candidate.sheet_state == "visible":
                    ws = candidate

            if ws is None:
                # Fallback to active or first visible worksheet
                active = wb.active
                if active and active.sheet_state == "visible":
                    ws = active
                else:
                    for s in wb.worksheets:
                        if s.sheet_state == "visible":
                            ws = s
                            break

            if ws is None:
                raise ValueError(
                    f"No visible worksheet found in '{bom_file.file_path}' "
                    f"matching sheet '{bom_file.sheet_name}'."
                )

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_list = list(row)

                if all(v is None or str(v).strip() == "" for v in row_list):
                    continue

                def get_val(col_idx: Optional[int], default="") -> str:
                    if col_idx is None or col_idx >= len(row_list):
                        return default
                    val = row_list[col_idx]
                    if val is None:
                        return default
                    return str(val).strip()

                def get_quantity_val(col_idx: Optional[int], default=""):
                    raw = get_val(col_idx, "")
                    if raw == "":
                        return default
                    try:
                        return parse_positive_integer_quantity(raw)
                    except ValueError:
                        # Preserve the original value so aggregation can
                        # skip it with an explicit warning.
                        return raw

                mpn = clean_mpn_value(get_val(mapping.mpn))
                quantity = get_quantity_val(mapping.quantity)

                board_val = get_val(mapping.board_identifier)
                if board_val:
                    board_name = f"Board {board_val}"
                else:
                    board_name = bom_file.board_name

                items.append(
                    BomItem(
                        source_file_name=source_file_name,
                        board_name=board_name,
                        comment=get_val(mapping.comment),
                        description=get_val(mapping.description),
                        designator=get_val(mapping.designator),
                        footprint=get_val(mapping.footprint),
                        quantity=quantity,
                        value=get_val(mapping.value),
                        manufacturer=get_val(mapping.manufacturer),
                        mpn=mpn,
                    )
                )
        finally:
            wb.close()
        return items

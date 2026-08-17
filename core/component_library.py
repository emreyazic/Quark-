"""Component library import reader and conflict resolution module.

Provides safe deduplication and conflict detection when importing Altium
component-library Excel workbooks. Ensures identical MPNs are merged safely
and different MPNs for the same internal code are flagged for explicit user resolution.
"""

from dataclasses import dataclass, field
from typing import Optional, List, Tuple, Dict, Set
from openpyxl import load_workbook

from core.mpn_utils import clean_mpn_value, normalize_mpn


INVALID_MPN_VALUES: Set[str] = {
    "*",
    "-",
    "N/A",
    "NA",
    "NONE",
    "TBD",
    "DNP",
    "DO NOT POPULATE",
    "NO POPULATE",
    "",
}


@dataclass
class LibraryRow:
    """Represents a single raw row read from a component library workbook."""
    row_number: int
    internal_code: str
    mpn: str
    normalized_mpn: str


@dataclass
class ConflictItem:
    """Represents an internal code with conflicting MPNs in file or with DB."""
    internal_code: str
    conflict_type: str  # "FILE_CONFLICT" or "DB_CONFLICT"
    candidate_mpns: List[str]  # Distinct raw MPNs from file
    row_numbers: Dict[str, List[int]]  # mpn -> list of Excel row numbers
    existing_db_mpn: Optional[str] = None  # Existing MPN in database if any


def _normalize_header(value) -> str:
    """Normalize Excel column headers for matching."""
    return " ".join(str(value or "").strip().upper().split())


def read_component_library_file(file_path: str) -> Tuple[List[LibraryRow], int]:
    """Read valid component rows from a library Excel file.

    Args:
        file_path: Path to the component library Excel workbook.

    Returns:
        tuple of (valid_rows, invalid_or_empty_skipped_count)
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = (
            workbook["Components"]
            if "Components" in workbook.sheetnames
            else workbook.active
        )
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("The selected file is empty.")

        header_map = {_normalize_header(v): i for i, v in enumerate(headers)}
        internal_index = header_map.get(
            "LIBRARYREFERENCE", header_map.get("COMMENT")
        )
        mpn_index = header_map.get("MANUFACTURER PART NUMBER")

        if internal_index is None or mpn_index is None:
            raise ValueError(
                "Required columns were not found. Expected LIBRARYREFERENCE "
                "(or COMMENT) and MANUFACTURER PART NUMBER."
            )

        valid_rows: List[LibraryRow] = []
        skipped = 0

        for row_number, row in enumerate(rows, start=2):
            if row_number is None or row is None:
                continue
            if internal_index >= len(row) or mpn_index >= len(row):
                skipped += 1
                continue

            raw_code = str(row[internal_index] or "").strip()
            raw_mpn = clean_mpn_value(row[mpn_index])
            norm_mpn = normalize_mpn(raw_mpn)

            if not raw_code or raw_mpn.upper() in INVALID_MPN_VALUES or not norm_mpn:
                skipped += 1
                continue

            valid_rows.append(
                LibraryRow(
                    row_number=row_number,
                    internal_code=raw_code,
                    mpn=raw_mpn,
                    normalized_mpn=norm_mpn,
                )
            )

        return valid_rows, skipped
    finally:
        workbook.close()


def detect_library_conflicts(
    raw_rows: List[LibraryRow],
    db_manager=None,
) -> Tuple[List[Tuple[str, str]], List[ConflictItem], int]:
    """Analyze library rows for internal code duplicates and conflicts.

    - Merges identical internal_code + same normalized MPN safely.
    - Flags internal_codes with multiple different MPNs as FILE_CONFLICT.
    - Flags internal_codes whose file MPN differs from existing DB mapping as DB_CONFLICT.

    Args:
        raw_rows: List of LibraryRow items parsed from file.
        db_manager: Optional DatabaseManager instance to check against existing mappings.

    Returns:
        tuple of (clean_components, conflicts, total_skipped_count)
        - clean_components: list of (internal_code, mpn) ready for import.
        - conflicts: list of ConflictItem requiring user review/resolution.
        - total_skipped_count: count of duplicate rows safely merged.
    """
    grouped_by_code: Dict[str, List[LibraryRow]] = {}
    for row in raw_rows:
        grouped_by_code.setdefault(row.internal_code, []).append(row)

    clean_components: List[Tuple[str, str]] = []
    conflicts: List[ConflictItem] = []
    merged_duplicates = 0

    # Fetch existing DB mappings if db_manager provided
    existing_db_mappings: Dict[str, str] = {}
    if db_manager is not None:
        try:
            for mapping in db_manager.get_all_internal_mappings():
                code = mapping.get("comment_code", "").strip()
                db_mpn = mapping.get("mpn", "").strip()
                if code and db_mpn:
                    existing_db_mappings[code] = db_mpn
        except Exception:
            existing_db_mappings = {}

    for internal_code, rows in grouped_by_code.items():
        # Group by normalized MPN
        unique_norm_mpns: Dict[str, Tuple[str, List[int]]] = {}
        for r in rows:
            if r.normalized_mpn not in unique_norm_mpns:
                unique_norm_mpns[r.normalized_mpn] = (r.mpn, [r.row_number])
            else:
                # Same internal code + same normalized MPN -> duplicate merged!
                unique_norm_mpns[r.normalized_mpn][1].append(r.row_number)
                merged_duplicates += 1

        db_existing_mpn = existing_db_mappings.get(internal_code)

        if len(unique_norm_mpns) == 1:
            # Single distinct MPN in file for this internal code
            norm_mpn, (raw_mpn, row_nums) = next(iter(unique_norm_mpns.items()))

            if db_existing_mpn:
                norm_db_mpn = normalize_mpn(db_existing_mpn)
                if norm_db_mpn != norm_mpn:
                    # Conflict with existing database mapping!
                    conflicts.append(
                        ConflictItem(
                            internal_code=internal_code,
                            conflict_type="DB_CONFLICT",
                            candidate_mpns=[raw_mpn],
                            row_numbers={raw_mpn: row_nums},
                            existing_db_mpn=db_existing_mpn,
                        )
                    )
                    continue

            # No conflict
            clean_components.append((internal_code, raw_mpn))
        else:
            # Multiple different MPNs in file for the same internal code!
            candidates = [raw_mpn for raw_mpn, _ in unique_norm_mpns.values()]
            row_map = {
                raw_mpn: row_nums
                for raw_mpn, row_nums in unique_norm_mpns.values()
            }
            conflicts.append(
                ConflictItem(
                    internal_code=internal_code,
                    conflict_type="FILE_CONFLICT",
                    candidate_mpns=candidates,
                    row_numbers=row_map,
                    existing_db_mpn=db_existing_mpn,
                )
            )

    return clean_components, conflicts, merged_duplicates

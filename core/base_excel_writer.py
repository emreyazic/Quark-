import re
import openpyxl
from datetime import datetime
from typing import Dict, Any, Optional, Union, Tuple
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from models.bom_item import BomItem
from core.mpn_utils import (
    parse_positive_integer_quantity,
    select_unit_price,
    select_digikey_price,
    get_safety_surplus,
    is_resistor_or_capacitor,
)


def add_refresh_changes_sheet(workbook, items: list[BomItem]) -> None:
    """Add a supplier price/stock delta report for the latest refresh."""
    sheet_name = "Refresh Changes"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    headers = [
        "MPN",
        "Supplier",
        "Previous Part Number",
        "Current Part Number",
        "Previous Stock",
        "Current Stock",
        "Stock Change",
        "Previous Unit Price",
        "Current Unit Price",
        "Unit Price Change",
        "Previous Observation",
        "Current Observation",
        "Previous Result",
        "Current Result",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    changes = [change for item in items for change in item.supplier_changes]
    if not changes:
        ws.append(["No supplier price or stock changes in the latest refresh."])
    else:
        for change in changes:
            previous_time = change.get("previous_observed_at")
            current_time = change.get("current_observed_at")
            ws.append([
                change.get("mpn", ""),
                change.get("supplier", ""),
                change.get("previous_part_number", ""),
                change.get("current_part_number", ""),
                change.get("previous_stock"),
                change.get("current_stock"),
                change.get("stock_change"),
                change.get("previous_unit_price"),
                change.get("current_unit_price"),
                change.get("unit_price_change"),
                datetime.fromtimestamp(previous_time).isoformat(sep=" ", timespec="seconds")
                if previous_time else "",
                datetime.fromtimestamp(current_time).isoformat(sep=" ", timespec="seconds")
                if current_time else "",
                change.get("previous_observation_type", ""),
                change.get("current_observation_type", ""),
            ])
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"0.000000'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        ws.column_dimensions[column[0].column_letter].width = width

class BaseExcelWriter:
    """Base class providing shared logic for Excel writers (Project and Workspace)."""
    
    def __init__(self, pricing_mode: str = "unit"):
        self.pricing_mode = pricing_mode
        self.wb = openpyxl.Workbook()
        if self.wb.sheetnames:
            active_sheet = self.wb.active
            assert active_sheet is not None
            self.wb.remove(active_sheet)

        self.header_font = Font(bold=True)
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.fill_green = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
        self.fill_yellow = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
        self.fill_red = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
        self.wrap_alignment = Alignment(vertical="center", wrap_text=True)

    def _fmt_qty(self, val: Union[int, float, str]) -> str:
        try:
            f = float(val)
            if f.is_integer():
                return str(int(f))
            return str(f)
        except (ValueError, TypeError):
            return str(val)

    def _fmt_qty_pair(self, per_board_qty: Union[int, float, str], total_qty: Union[int, float, str]) -> str:
        return f"{self._fmt_qty(per_board_qty)} / {self._fmt_qty(total_qty)}"

    def _format_usage_quantities(self, usages) -> str:
        return "\n".join(
            self._fmt_qty_pair(usage.bom_line_quantity, usage.total_quantity)
            for usage in usages
        )

    def _apply_narrow_columns(self, ws, header_row: int = 1):
        headers = [str(cell.value) for cell in ws[header_row]]
        narrow_targets = {
            "Quantity (Per Board / Total)": 18,
            "Quantity": 15,
            "Description": 35,
            "Designator": 25
        }
        for col_idx, header in enumerate(headers, 1):
            if header in narrow_targets:
                ws.column_dimensions[get_column_letter(col_idx)].width = narrow_targets[header]

    def _auto_fit_columns(self, ws, min_width: int = 8, max_width: int = 80, padding: int = 6) -> None:
        """Set worksheet column widths from the longest visible line in each column."""
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                value = row[0].value
                if value is None:
                    continue

                lines = str(value).splitlines() or [""]
                max_length = max(max_length, *(len(line) for line in lines))

            if max_length:
                width = min(max(max_length + padding, min_width), max_width)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _sanitize_sheet_name(self, name: str) -> str:
        s = re.sub(r'[\\\\/?*\\[\\]:]', '', name).strip()
        return s[:31]

    def _safe_sheet_name(self, base_name: str, used_names: set[str]) -> str:
        """Generates a safe and unique sheet name from the base_name."""
        s = self._sanitize_sheet_name(base_name)
        if not s:
            s = "Sheet"
        sheet_name = s
        counter = 2
        while sheet_name in used_names:
            suffix = f" {counter}"
            sheet_name = s[:31 - len(suffix)] + suffix
            counter += 1
        used_names.add(sheet_name)
        return sheet_name

    def _format_currency(self, cell):
        cell.number_format = '"$"#,##0.0000'

    def _format_supplier_price_cells(self, ws, row: int, jlc_col: int, digikey_col: int, item: BomItem) -> None:
        """Format separate supplier prices and highlight the cheaper option."""
        jlc_cell = ws.cell(row=row, column=jlc_col)
        digikey_cell = ws.cell(row=row, column=digikey_col)
        for cell in (jlc_cell, digikey_cell):
            if isinstance(cell.value, (int, float)):
                self._format_currency(cell)
        if item.unit_price is not None and item.digikey_unit_price is not None:
            if item.unit_price < item.digikey_unit_price:
                jlc_cell.fill, digikey_cell.fill = self.fill_green, self.fill_red
            elif item.digikey_unit_price < item.unit_price:
                jlc_cell.fill, digikey_cell.fill = self.fill_red, self.fill_green
        elif item.unit_price is not None:
            jlc_cell.fill = self.fill_green
        elif item.digikey_unit_price is not None:
            digikey_cell.fill = self.fill_green

    def _component_price_values(self, item: BomItem, quantity=None):
        """Return display quantity, supplier unit prices, and optional totals."""
        raw_quantity = quantity if quantity is not None else item.pricing_quantity
        try:
            pricing_quantity = parse_positive_integer_quantity(raw_quantity)
        except (ValueError, TypeError):
            return 0, None, None, None, None
        use_breaks = self.pricing_mode == "project"
        j_price = select_unit_price(
            item.jlcpcb_price_breaks_raw,
            pricing_quantity,
            use_quantity_breaks=use_breaks,
        ) if item.jlcpcb_price_breaks_raw else item.unit_price
        d_price = select_digikey_price(
            item.digikey_price_breaks,
            pricing_quantity,
            use_quantity_breaks=use_breaks,
        ) if item.digikey_price_breaks else item.digikey_unit_price
        return (
            pricing_quantity,
            j_price,
            d_price,
            pricing_quantity * j_price if j_price is not None else None,
            pricing_quantity * d_price if d_price is not None else None,
        )

    def _selected_supplier_price(
        self,
        item: BomItem,
        j_price: Optional[float],
        d_price: Optional[float],
        required_quantity: Optional[Union[int, float]] = None,
    ) -> Tuple[str, Optional[float]]:
        """Choose the price used by the mixed-sourcing cost calculation.
        
        Requires a valid supplier part number, valid unit price, and
        sufficient available stock (stock is NOT assumed sufficient if None or < required_quantity).
        """
        qty = required_quantity if required_quantity is not None else item.pricing_quantity
        try:
            qty_int = parse_positive_integer_quantity(qty)
        except (ValueError, TypeError):
            qty_int = 1

        j_purchasable = (
            self._is_jlcpcb_usable(item)
            and bool(item.jlcpcb_part_number)
            and j_price is not None
            and item.available_stock_qty is not None
            and item.available_stock_qty >= qty_int
            and qty_int > 0
        )
        if j_purchasable:
            return "JLCPCB", j_price

        d_purchasable = (
            bool(item.digikey_part_number)
            and item.digikey_status == "found"
            and d_price is not None
            and item.digikey_stock_qty is not None
            and item.digikey_stock_qty >= qty_int
            and qty_int > 0
        )
        if d_purchasable:
            return "DigiKey fallback", d_price

        if (
            j_price is not None
            or d_price is not None
            or bool(item.jlcpcb_part_number)
            or bool(item.digikey_part_number)
        ):
            return "Shortage / Unavailable", None

        return "Unpriced", None

    def _add_price_totals_box(
        self,
        ws,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        build_quantity: int,
        total_label: str = "Project Total",
        missing_price_count: Optional[int] = None,
    ) -> None:
        """Add totals plus an explicit warning when one or more costs are unknown."""
        if data_end_row < data_start_row:
            return

        headers = [str(cell.value or "") for cell in ws[header_row]]
        try:
            jlc_total_col = headers.index("JLCPCB Total Price") + 1
            digikey_total_col = headers.index("DigiKey Total Price") + 1
        except ValueError:
            return
        if "Purchase Quantity" in headers:
            label_col = headers.index("Purchase Quantity") + 1
        elif "Pricing Quantity" in headers:
            label_col = headers.index("Pricing Quantity") + 1
        elif "Pricing Pool Quantity" in headers:
            label_col = headers.index("Pricing Pool Quantity") + 1
        else:
            return

        total_row = data_end_row + 2
        per_board_row = total_row + 1
        build_quantity = max(int(build_quantity or 1), 1)

        ws.cell(row=total_row, column=label_col, value=total_label)
        ws.cell(row=per_board_row, column=label_col, value="Cost Per Board Set")
        for row in (total_row, per_board_row):
            ws.cell(row=row, column=label_col).font = self.header_font
            ws.cell(row=row, column=label_col).fill = self.fill_yellow

        total_columns = [jlc_total_col, digikey_total_col]
        if "Mixed Sourcing Total Price" in headers:
            total_columns.append(headers.index("Mixed Sourcing Total Price") + 1)

        for total_col in total_columns:
            col_letter = get_column_letter(total_col)
            total_cell = ws.cell(
                row=total_row,
                column=total_col,
                value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})",
            )
            per_board_cell = ws.cell(
                row=per_board_row,
                column=total_col,
                value=f"={col_letter}{total_row}/{build_quantity}",
            )
            for cell in (total_cell, per_board_cell):
                cell.font = self.header_font
                cell.fill = self.fill_yellow
                self._format_currency(cell)

        if missing_price_count is not None:
            missing_price_count = max(int(missing_price_count), 0)
            missing_row = per_board_row + 1
            status_row = missing_row + 1
            ws.cell(row=missing_row, column=label_col, value="Missing Price Count")
            ws.cell(row=missing_row, column=jlc_total_col, value=missing_price_count)
            ws.cell(row=status_row, column=label_col, value="Cost Status")
            ws.cell(
                row=status_row,
                column=jlc_total_col,
                value="COMPLETE" if missing_price_count == 0 else "INCOMPLETE",
            )
            warning_fill = self.fill_green if missing_price_count == 0 else self.fill_red
            for row in (missing_row, status_row):
                for col in (label_col, jlc_total_col):
                    cell = ws.cell(row=row, column=col)
                    cell.font = self.header_font
                    cell.fill = warning_fill

    def _add_board_cost_summary(
        self,
        ws,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        build_quantity: int,
        missing_price_count: int = 0,
    ) -> None:
        """Place the mixed-sourcing card cost prominently at the top of a board sheet."""
        if data_end_row < data_start_row:
            return

        headers = [str(cell.value or "") for cell in ws[header_row]]
        if "Mixed Sourcing Total Price" not in headers:
            return

        mixed_col = headers.index("Mixed Sourcing Total Price") + 1
        mixed_letter = get_column_letter(mixed_col)
        build_quantity = max(int(build_quantity or 1), 1)

        ws["D1"] = "Card Production Cost (Mixed Sourcing):"
        ws["E1"] = f"=SUM({mixed_letter}{data_start_row}:{mixed_letter}{data_end_row})"
        ws["D2"] = "Cost Per Card:"
        ws["E2"] = f"=E1/{build_quantity}"
        ws["D3"] = "Missing Price Count:"
        ws["E3"] = max(int(missing_price_count), 0)
        ws["D4"] = "Cost Status:"
        ws["E4"] = "COMPLETE" if missing_price_count == 0 else "INCOMPLETE"
        for cell_ref in ("D1", "E1", "D2", "E2"):
            ws[cell_ref].font = self.header_font
            ws[cell_ref].fill = self.fill_yellow
        warning_fill = self.fill_green if missing_price_count == 0 else self.fill_red
        for cell_ref in ("D3", "E3", "D4", "E4"):
            ws[cell_ref].font = self.header_font
            ws[cell_ref].fill = warning_fill
        self._format_currency(ws["E1"])
        self._format_currency(ws["E2"])

    def _add_supplier_stock_sheet(self, items: list[BomItem]) -> None:
        """Write one JLCPCB row and one DigiKey row per component."""
        sheet_name = "Supplier Stock"
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        ws = self.wb.create_sheet(sheet_name)
        headers = [
            "MPN", "Design Item ID", "Supplier", "Supplier Part Number",
            "Stock", "Unit Price", "Production Required Quantity", "Safety Surplus",
            "Purchase Quantity", "Total Price", "Required Stock", "Status",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        for item in items:
            purchase_qty, j_price, d_price, j_total, d_total = self._component_price_values(item)
            prod_qty = item.production_quantity_val
            surplus = item.safety_surplus
            supplier_rows = [
                (
                    "JLCPCB",
                    item.jlcpcb_part_number,
                    item.available_stock_qty,
                    j_price,
                    j_total,
                    (
                        f"API Error: {item.jlcpcb_error}"
                        if item.jlcpcb_status == "error"
                        else item.jlcpcb_status.replace("_", " ").title()
                    ),
                ),
                (
                    "DigiKey",
                    item.digikey_part_number,
                    item.digikey_stock_qty,
                    d_price,
                    d_total,
                    (
                        f"API Error: {item.digikey_error}"
                        if item.digikey_status == "error"
                        else item.digikey_status.replace("_", " ").title()
                    ),
                ),
            ]
            for supplier, part_number, stock, unit_price, total_price, status in supplier_rows:
                ws.append([
                    item.mpn,
                    item.comment,
                    supplier,
                    part_number or "-",
                    stock if stock is not None else "-",
                    unit_price,
                    prod_qty,
                    surplus,
                    purchase_qty,
                    total_price,
                    item.required_stock,
                    status or "",
                ])
                if isinstance(unit_price, (int, float)):
                    self._format_currency(ws.cell(row=ws.max_row, column=6))
                if isinstance(total_price, (int, float)):
                    self._format_currency(ws.cell(row=ws.max_row, column=10))

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws, max_width=40)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 22)
        ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 24)

    def _get_status_fill(self, status_text: str, has_jlcpcb_part: bool = False):
        """Returns the appropriate PatternFill based on status text and JLCPCB part presence."""
        val = str(status_text or "")
        status_lower = val.lower()
        
        # Failure checks must happen before success checks
        if "not found" in status_lower or "error" in status_lower or "no exact" in status_lower or "insufficient" in status_lower:
            return self.fill_red
            
        # Success checks
        if "✅" in val or "found" in status_lower or (val == "" and has_jlcpcb_part):
            return self.fill_green
            
        # Warning/Unknown checks
        if val != "":
            return self.fill_yellow
            
        return None

    def _is_jlcpcb_usable(self, item: BomItem) -> bool:
        """Determines whether JLCPCB pricing is usable for this component."""
        status_lower = str(item.status or "").lower()
        if item.jlcpcb_status in ("not_found", "error", "mismatch"):
            return False
        if not item.jlcpcb_part_number:
            return False
        if "not found" in status_lower:
            return False
        if "no exact" in status_lower or "mismatch" in status_lower:
            return False
        if "error" in status_lower:
            return False
        
        return True

    def _get_pricing_for_component_item(self, item: BomItem, multiplied_qty: Union[int, float]) -> Dict[str, Any]:
        """Calculate JLCPCB, DigiKey fallback, and DigiKey-only costs based on stock availability and purchase quantity."""
        try:
            prod_qty = parse_positive_integer_quantity(multiplied_qty)
        except (ValueError, TypeError):
            prod_qty = 0
        surplus = getattr(item, "safety_surplus", 0) or get_safety_surplus(item)
        purchase_quantity = prod_qty + surplus if prod_qty > 0 else 0

        use_breaks = self.pricing_mode == "project"
        j_price = select_unit_price(
            item.jlcpcb_price_breaks_raw,
            purchase_quantity,
            use_quantity_breaks=use_breaks,
        ) if item.jlcpcb_price_breaks_raw else item.unit_price
        d_price = select_digikey_price(
            item.digikey_price_breaks,
            purchase_quantity,
            use_quantity_breaks=use_breaks,
        ) if item.digikey_price_breaks else item.digikey_unit_price

        if purchase_quantity == 0:
            j_price = None
            d_price = None
        elif not self._is_jlcpcb_usable(item):
            j_price = None

        j_cost = None
        rem_dk_cost = None
        all_dk_cost = None
        selected_price = None
        selected_source = "Unpriced"
        combined_cost = None

        # Sourcing purchasability criteria: checks stock >= purchase_quantity
        j_purchasable = (
            self._is_jlcpcb_usable(item)
            and bool(item.jlcpcb_part_number)
            and j_price is not None
            and item.available_stock_qty is not None
            and item.available_stock_qty >= purchase_quantity
            and purchase_quantity > 0
        )
        d_purchasable = (
            bool(item.digikey_part_number)
            and item.digikey_status == "found"
            and d_price is not None
            and item.digikey_stock_qty is not None
            and item.digikey_stock_qty >= purchase_quantity
            and purchase_quantity > 0
        )

        if d_purchasable and d_price is not None:
            all_dk_cost = purchase_quantity * d_price

        if j_purchasable and j_price is not None:
            selected_source = "JLCPCB"
            selected_price = j_price
            j_cost = purchase_quantity * j_price
            combined_cost = j_cost
        elif d_purchasable and d_price is not None:
            selected_source = "DigiKey fallback"
            selected_price = d_price
            rem_dk_cost = purchase_quantity * d_price
            combined_cost = rem_dk_cost
        elif (
            j_price is not None
            or d_price is not None
            or bool(item.jlcpcb_part_number)
            or bool(item.digikey_part_number)
        ):
            selected_source = "Shortage / Unavailable"
            selected_price = None
            combined_cost = None

        return {
            "selected_source": selected_source,
            "selected_price": selected_price,
            "jlcpcb_cost": j_cost,
            "remaining_digikey_cost": rem_dk_cost,
            "combined_cost": combined_cost,
            "digikey_only_cost": all_dk_cost,
            "j_price": j_price,
            "d_price": d_price,
            "production_quantity": prod_qty,
            "safety_surplus": surplus,
            "purchase_quantity": purchase_quantity,
        }

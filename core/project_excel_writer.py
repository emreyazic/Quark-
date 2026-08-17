import os
from typing import List, Dict, Any, Optional, Union
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import datetime

from models.bom_item import BomItem
from models.project import Project
from services.project_aggregation import ProjectAggregationResult, AggregatedComponent
from core.base_excel_writer import BaseExcelWriter, add_refresh_changes_sheet
from core.atomic_io import atomic_save_workbook


class ProjectExcelWriter(BaseExcelWriter):
    """Writes a project-aware multi-board BOM cost report to Excel."""

    def __init__(
        self,
        project: Project,
        aggregation_result: ProjectAggregationResult,
        enriched_items: List[BomItem],
        component_keys: List[str],
        build_multipliers: Optional[List[int]] = None,
        include_raw_board_sheets: bool = True,
        pricing_mode: str = "unit",
    ):
        if len(enriched_items) != len(component_keys):
            raise ValueError(
                f"Mismatch: {len(enriched_items)} enriched items vs {len(component_keys)} component keys."
            )

        aggregation_keys = {c.component_key for c in aggregation_result.components}
        enriched_keys = set(component_keys)
        
        if len(enriched_keys) != len(component_keys):
            raise ValueError("Duplicate component keys found in component_keys list.")

        missing = aggregation_keys - enriched_keys
        extra = enriched_keys - aggregation_keys

        if missing or extra:
            raise ValueError(
                f"Component key mismatch. Missing: {missing}, Extra: {extra}"
            )

        self.project = project
        self.aggregation_result = aggregation_result
        self.enriched_items = enriched_items
        self.component_keys = component_keys
        self.build_multipliers = build_multipliers or [1, 5, 10, 50, 100]
        self.include_raw_board_sheets = include_raw_board_sheets

        self.enriched_by_key: Dict[str, BomItem] = {}
        for k, item in zip(self.component_keys, self.enriched_items):
            self.enriched_by_key[k] = item

        self.component_by_key: Dict[str, AggregatedComponent] = {
            c.component_key: c for c in self.aggregation_result.components
        }

        BaseExcelWriter.__init__(self, pricing_mode=pricing_mode)

    def _get_pricing_for_component(self, comp_key: str, multiplied_qty: Union[int, float]) -> Dict[str, Any]:
        """Calculate JLCPCB, DigiKey fallback, and DigiKey-only costs."""
        item = self.enriched_by_key[comp_key]
        return self._get_pricing_for_component_item(item, multiplied_qty)

    def write(self, output_path: str):
        self._add_master_summary_sheet()
        self._add_aggregated_components_sheet()
        self._add_per_board_sheets()
        if self.include_raw_board_sheets:
            self._add_raw_board_sheets()
        self._add_supplier_stock_sheet(self.enriched_items)
        add_refresh_changes_sheet(self.wb, self.enriched_items)
            
        atomic_save_workbook(self.wb, output_path)

    def _add_master_summary_sheet(self):
        ws = self.wb.create_sheet("Master Summary")

        # Project Info
        ws.append(["Project Info"])
        ws["A1"].font = Font(bold=True, size=14)
        
        ws.append(["Project Name:", self.project.project_name])
        ws.append(["Generated At:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        actual_board_names = {
            usage.board_name
            for component in self.aggregation_result.components
            for usage in component.usages
        }
        ws.append(["Number of Boards:", len(actual_board_names)])
        ws.append(["Number of BOM Files:", len(self.project.board_items)])
        ws.append(["Unique Components:", len(self.component_keys)])
        ws.append(["Skipped Rows:", self.aggregation_result.skipped_count])
        ws.append(["Build Multipliers:", ", ".join(map(str, self.build_multipliers))])
        ws.append([
            "Mixed Sourcing Logic:",
            "Use JLCPCB when its code and price exist; use DigiKey only as fallback. "
            "The same component is never counted twice.",
        ])
        ws.append([])

        # Board Composition
        ws.append(["Board Composition"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=14)
        
        ws.append(["Board Name", "Board Quantity", "Source BOM File", "Parsed Item Count"])
        for cell in ws[ws.max_row]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        for board in self.project.board_items:
            ws.append([
                board.board_name,
                self._fmt_qty(board.board_quantity),
                os.path.basename(board.file_path),
                len(board.bom_items) if board.bom_items else 0
            ])
        ws.append([])

        # Cost Summary
        ws.append(["Cost Summary"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=14)
        
        ws.append([
            "Build Qty", "JLCPCB Cost", "DigiKey Fallback Cost",
            "Mixed Sourcing Total", "DigiKey-Only Total",
            "Missing Price Count", "Cost Status",
        ])
        for cell in ws[ws.max_row]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        # Calculate totals per multiplier
        for m in self.build_multipliers:
            tot_jlc = 0.0
            tot_rem_dk = 0.0
            tot_comb = 0.0
            tot_dk_only = 0.0
            missing_price_count = 0
            
            for comp_key, comp in self.component_by_key.items():
                mult_qty = comp.total_quantity * m
                pricing = self._get_pricing_for_component(comp_key, mult_qty)
                
                if pricing["jlcpcb_cost"] is not None: tot_jlc += pricing["jlcpcb_cost"]
                if pricing["remaining_digikey_cost"] is not None: tot_rem_dk += pricing["remaining_digikey_cost"]
                if pricing["combined_cost"] is not None: tot_comb += pricing["combined_cost"]
                else: missing_price_count += 1
                if pricing["digikey_only_cost"] is not None: tot_dk_only += pricing["digikey_only_cost"]

            cost_status = "COMPLETE" if missing_price_count == 0 else "INCOMPLETE"
            row = [
                f"{m}x", tot_jlc, tot_rem_dk, tot_comb, tot_dk_only,
                missing_price_count, cost_status,
            ]
            ws.append(row)
            for c_idx in range(2, 6):
                self._format_currency(ws.cell(row=ws.max_row, column=c_idx))
            for c_idx in (6, 7):
                ws.cell(row=ws.max_row, column=c_idx).fill = (
                    self.fill_green if missing_price_count == 0 else self.fill_red
                )

        ws.append([])

        # Warning Summary
        if self.aggregation_result.warnings:
            ws.append(["Aggregation Warnings"])
            ws[f"A{ws.max_row}"].font = Font(bold=True, size=14, color="FF0000")
            for w in self.aggregation_result.warnings:
                ws.append([w])
                ws.cell(row=ws.max_row, column=1).font = Font(color="FF0000")
                
        self._auto_fit_columns(ws)

    def _add_aggregated_components_sheet(self):
        ws = self.wb.create_sheet("Aggregated Components")
        
        headers = [
            "Board Name", "Description", "Designator", "Quantity (Per Board / Total)", 
            "Value", "Manufacturer", "MPN", "Design Item ID", "JLCPCB Part Number", "DigiKey Part Number", "JLCPCB Stock", "DigiKey Stock",
            "JLCPCB Unit Price", "DigiKey Unit Price", "Pricing Quantity", "JLCPCB Total Price", "DigiKey Total Price", "Status"
        ]

        ws.append(headers)
        data_start_row = 2
        for cell in ws[1]:
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            
        for comp_key, comp in self.component_by_key.items():
            enriched = self.enriched_by_key[comp_key]
            
            # Aggregate designators
            all_desigs = []
            for u in comp.usages:
                if u.bom_item.designator:
                    all_desigs.append(str(u.bom_item.designator))
            desigs_str = ", ".join(all_desigs)[:100] # Cap length
            if len(", ".join(all_desigs)) > 100:
                desigs_str += "..."
                
            # Aggregate source boards
            boards_usage = []
            for u in comp.usages:
                boards_usage.append(f"{u.board_name}: {self._fmt_qty(u.total_quantity)}")
            boards_str = "\n".join(boards_usage)
            quantities_str = self._format_usage_quantities(comp.usages)
            pricing_quantity, j_price, d_price, j_total, d_total = self._component_price_values(enriched)

            row = [
                boards_str,
                enriched.description,
                desigs_str,
                quantities_str,
                enriched.value,
                enriched.manufacturer,
                enriched.mpn,
                enriched.comment,
                enriched.jlcpcb_part_number,
                enriched.digikey_part_number,
                enriched.available_stock_qty if enriched.available_stock_qty is not None else "-",
                enriched.digikey_stock_qty if enriched.digikey_stock_qty is not None else "-",
                j_price,
                d_price,
                pricing_quantity,
                j_total,
                d_total,
                enriched.status
            ]
                
            ws.append(row)
            ws.cell(row=ws.max_row, column=1).alignment = self.wrap_alignment
            ws.cell(row=ws.max_row, column=4).alignment = self.wrap_alignment
            
            # Only status coloring left since we removed multi-column pricing
            self._format_supplier_price_cells(ws, ws.max_row, 13, 14, enriched)
            for total_col in (16, 17):
                if isinstance(ws.cell(row=ws.max_row, column=total_col).value, (int, float)):
                    self._format_currency(ws.cell(row=ws.max_row, column=total_col))
            cell = ws.cell(row=ws.max_row, column=18)
            fill = self._get_status_fill(enriched.status, bool(enriched.jlcpcb_part_number))
            if fill:
                cell.fill = fill

        data_end_row = ws.max_row
        missing_price_count = sum(
            1
            for enriched in self.enriched_by_key.values()
            if self._selected_supplier_price(
                enriched,
                self._component_price_values(enriched)[1],
                self._component_price_values(enriched)[2],
            )[1] is None
        )
        self._add_price_totals_box(
            ws, 1, data_start_row, data_end_row, self.build_multipliers[0],
            missing_price_count=missing_price_count,
        )
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end_row}"
        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws)

    def _add_per_board_sheets(self):
        used_names = {"Master Summary", "Aggregated Components"}
        
        for board in self.project.board_items:
            base_name = self._sanitize_sheet_name(board.board_name)
            if not base_name:
                base_name = "Board"
                
            sheet_name = base_name
            counter = 2
            while sheet_name in used_names:
                suffix = f" {counter}"
                sheet_name = base_name[:31 - len(suffix)] + suffix
                counter += 1
            used_names.add(sheet_name)
            
            ws = self.wb.create_sheet(sheet_name)
            
            # Header info
            ws.append(["Project Name:", self.project.project_name])
            ws.append(["Board Name:", board.board_name])
            ws.append(["Board Quantity:", self._fmt_qty(board.board_quantity)])
            ws.append(["Source BOM File:", os.path.basename(board.file_path)])
            ws.append([])
            
            for row_idx in range(1, 5):
                ws.cell(row=row_idx, column=1).font = self.header_font

            headers = [
                "Board Name", "Description", "Designator", "Quantity (Per Board / Total)", 
                "Value", "Manufacturer", "MPN", "Design Item ID", "JLCPCB Part Number", "DigiKey Part Number", "JLCPCB Stock", "DigiKey Stock",
                "JLCPCB Unit Price", "DigiKey Unit Price", "Pricing Quantity", "JLCPCB Total Price", "DigiKey Total Price",
                "Selected Supplier", "Mixed Sourcing Total Price", "Status"
            ]
            ws.append(headers)
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                
            # We must map usages to this specific board.
            # A board's component is found in the AggregatedComponent.usages.
            missing_price_count = 0
            for comp_key, comp in self.component_by_key.items():
                enriched = self.enriched_by_key[comp_key]
                
                # Find all usages of this component on THIS board
                board_usages = [u for u in comp.usages if u.board_file_path == board.file_path]
                if not board_usages:
                    continue
                display_quantity = sum(u.total_quantity for u in board_usages) * self.build_multipliers[0]
                pricing_quantity, j_price, d_price, j_total, d_total = self._component_price_values(enriched, display_quantity)
                selected_source, selected_price = self._selected_supplier_price(
                    enriched, j_price, d_price
                )
                mixed_total = display_quantity * selected_price if selected_price is not None else None
                if selected_price is None:
                    missing_price_count += 1

                for usage_index, usage in enumerate(board_usages):
                    row = [
                        board.board_name,
                        enriched.description,
                        usage.bom_item.designator,
                        f"{self._fmt_qty(usage.bom_line_quantity)} / {self._fmt_qty(usage.total_quantity)}",
                        enriched.value,
                        enriched.manufacturer,
                        enriched.mpn,
                        usage.bom_item.comment,
                        enriched.jlcpcb_part_number,
                        enriched.digikey_part_number,
                        enriched.available_stock_qty if enriched.available_stock_qty is not None else "-",
                        enriched.digikey_stock_qty if enriched.digikey_stock_qty is not None else "-",
                        j_price,
                        d_price,
                        pricing_quantity,
                        j_total if usage_index == 0 else None,
                        d_total if usage_index == 0 else None,
                        selected_source if usage_index == 0 else None,
                        mixed_total if usage_index == 0 else None,
                        enriched.status
                    ]
                    ws.append(row)
                    
                    # Status coloring
                    self._format_supplier_price_cells(ws, ws.max_row, 13, 14, enriched)
                    for total_col in (16, 17):
                        if isinstance(ws.cell(row=ws.max_row, column=total_col).value, (int, float)):
                            self._format_currency(ws.cell(row=ws.max_row, column=total_col))
                    if isinstance(ws.cell(row=ws.max_row, column=19).value, (int, float)):
                        self._format_currency(ws.cell(row=ws.max_row, column=19))
                    cell = ws.cell(row=ws.max_row, column=20)
                    fill = self._get_status_fill(enriched.status, bool(enriched.jlcpcb_part_number))
                    if fill:
                        cell.fill = fill
                            
            data_start_row = header_row + 1
            data_end_row = ws.max_row
            self._add_price_totals_box(
                ws,
                header_row,
                data_start_row,
                data_end_row,
                self.build_multipliers[0],
                total_label="Board Total",
                missing_price_count=missing_price_count,
            )
            self._add_board_cost_summary(
                ws,
                header_row,
                data_start_row,
                data_end_row,
                board.board_quantity * self.build_multipliers[0],
                missing_price_count=missing_price_count,
            )
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{data_end_row}"
            ws.freeze_panes = f"A{header_row + 1}"
            self._auto_fit_columns(ws)

    def _add_raw_board_sheets(self):
        used_names = set(self.wb.sheetnames)
        
        for board in self.project.board_items:
            base_name = "Raw - " + self._sanitize_sheet_name(board.board_name)[:25]
            sheet_name = base_name
            counter = 2
            while sheet_name in used_names:
                suffix = f" {counter}"
                sheet_name = base_name[:31 - len(suffix)] + suffix
                counter += 1
            used_names.add(sheet_name)
            
            ws = self.wb.create_sheet(sheet_name)
            
            headers = BomItem.get_headers()
            ws.append(headers)
            qty_col_idx = headers.index("Quantity (Per Board / Total)")
            for cell in ws[1]:
                cell.font = self.header_font
                cell.alignment = self.header_alignment
                
            if board.bom_items:
                for item in board.bom_items:
                    row = item.to_row()
                    row[qty_col_idx] = self._fmt_qty(item.quantity)
                    ws.append(row)
                    
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            self._auto_fit_columns(ws)
            

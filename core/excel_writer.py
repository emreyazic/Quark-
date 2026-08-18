import os
from typing import List, Optional, cast
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.bom_item import BomItem
from core.atomic_io import atomic_save_workbook
from core.mpn_utils import parse_positive_integer_quantity
from core.base_excel_writer import add_refresh_changes_sheet


class ExcelWriter:
    """Writes enriched BOM data to an Excel file with strict formatting rules."""

    def __init__(self, items: List[BomItem], pricing_mode: str = "unit", build_multipliers: Optional[List[int]] = None):
        self.items = items
        self.pricing_mode = pricing_mode
        self.build_multipliers = build_multipliers or [1, 5, 10, 50, 100]
        self.wb = openpyxl.Workbook()
        self.ws = cast(Worksheet, self.wb.active)
        self.ws.title = "Enriched BOM"

        # Define fills specifically for the JLCPCB Part Number column
        self.fill_green = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
        self.fill_yellow = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
        self.fill_red = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")

    def write(self, output_path: str) -> None:
        """Write the BOM data to the given Excel file path."""
        headers = BomItem.get_headers()
        self.ws.append(headers)

        # Style headers
        header_font = Font(bold=True)
        for col_idx, cell in enumerate(self.ws[1], 1):
            cell.font = header_font
            self.ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Adjust specific column widths
        try:
            self.ws.column_dimensions[get_column_letter(headers.index("Description") + 1)].width = 40
            self.ws.column_dimensions[get_column_letter(headers.index("Designator") + 1)].width = 30
            self.ws.column_dimensions[get_column_letter(headers.index("JLCPCB Stock") + 1)].width = 16
            self.ws.column_dimensions[get_column_letter(headers.index("DigiKey Stock") + 1)].width = 16
            self.ws.column_dimensions[get_column_letter(headers.index("DigiKey Part Number") + 1)].width = 24
            self.ws.column_dimensions[get_column_letter(headers.index("JLCPCB Unit Price") + 1)].width = 20
            self.ws.column_dimensions[get_column_letter(headers.index("DigiKey Unit Price") + 1)].width = 20
            self.ws.column_dimensions[get_column_letter(headers.index("JLCPCB Total Price") + 1)].width = 20
            self.ws.column_dimensions[get_column_letter(headers.index("DigiKey Total Price") + 1)].width = 20
        except ValueError:
            pass

        jlcpcb_part_col_idx = headers.index("JLCPCB Part Number") + 1
        quantity_col_idx = headers.index("Quantity (Per Board / Total)") + 1
        jlcpcb_stock_col_idx = headers.index("JLCPCB Stock") + 1
        digikey_stock_col_idx = headers.index("DigiKey Stock") + 1
        jlcpcb_price_col_idx = headers.index("JLCPCB Unit Price") + 1
        digikey_price_col_idx = headers.index("DigiKey Unit Price") + 1
        jlcpcb_total_col_idx = headers.index("JLCPCB Total Price") + 1
        digikey_total_col_idx = headers.index("DigiKey Total Price") + 1

        for row_idx, item in enumerate(self.items, 2):
            row_data = item.to_row()

            for col_idx, val in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=val)

                # Alignment
                if col_idx in (quantity_col_idx, jlcpcb_part_col_idx, jlcpcb_stock_col_idx, digikey_stock_col_idx, jlcpcb_price_col_idx, digikey_price_col_idx):
                    cell.alignment = Alignment(horizontal="center")

                if col_idx in (jlcpcb_price_col_idx, digikey_price_col_idx, jlcpcb_total_col_idx, digikey_total_col_idx) and isinstance(val, (int, float)):
                    cell.number_format = '"$"0.00000'

                # Apply specific coloring ONLY to the JLCPCB Part Number cell
                if col_idx == jlcpcb_part_col_idx:
                    if "error" in item.status.lower() or "exception" in item.status.lower():
                        cell.fill = self.fill_red
                    elif item.status == "" and item.jlcpcb_part_number:
                        cell.fill = self.fill_green
                    else:
                        # Any other problem (missing MPN, RES, mismatch, insufficient stock) is yellow
                        cell.fill = self.fill_yellow

            jlc_price = item.unit_price
            digikey_price = item.digikey_unit_price
            if jlc_price is not None and digikey_price is not None:
                if jlc_price < digikey_price:
                    self.ws.cell(row=row_idx, column=jlcpcb_price_col_idx).fill = self.fill_green
                    self.ws.cell(row=row_idx, column=digikey_price_col_idx).fill = self.fill_red
                elif digikey_price < jlc_price:
                    self.ws.cell(row=row_idx, column=jlcpcb_price_col_idx).fill = self.fill_red
                    self.ws.cell(row=row_idx, column=digikey_price_col_idx).fill = self.fill_green
            elif jlc_price is not None:
                self.ws.cell(row=row_idx, column=jlcpcb_price_col_idx).fill = self.fill_green
            elif digikey_price is not None:
                self.ws.cell(row=row_idx, column=digikey_price_col_idx).fill = self.fill_green

        total_row = len(self.items) + 2
        # "Total Cost" must always sum extended line costs. Unit-price mode
        # controls which unit price is selected; it must not turn the project
        # total into a meaningless sum of per-unit prices.
        if "Purchase Quantity" in headers:
            label_col_idx = headers.index("Purchase Quantity") + 1
        elif "Pricing Quantity" in headers:
            label_col_idx = headers.index("Pricing Quantity") + 1
        else:
            label_col_idx = 1
        total_columns = (jlcpcb_total_col_idx, digikey_total_col_idx)
        self.ws.cell(row=total_row, column=label_col_idx, value="Total Cost")
        self.ws.cell(row=total_row, column=label_col_idx).font = header_font
        for price_col_idx in total_columns:
            price_col_letter = get_column_letter(price_col_idx)
            total_cell = self.ws.cell(
                row=total_row,
                column=price_col_idx,
                value=f"=SUM({price_col_letter}2:{price_col_letter}{total_row - 1})",
            )
            total_cell.font = header_font
            total_cell.number_format = '"$"0.00000'

        missing_price_count = sum(1 for item in self.items if self._is_price_missing(item))
        warning_col_idx = total_columns[0]
        self.ws.cell(row=total_row + 1, column=label_col_idx, value="Missing Price Count")
        self.ws.cell(row=total_row + 1, column=warning_col_idx, value=missing_price_count)
        self.ws.cell(row=total_row + 2, column=label_col_idx, value="Cost Status")
        self.ws.cell(
            row=total_row + 2,
            column=warning_col_idx,
            value="COMPLETE" if missing_price_count == 0 else "INCOMPLETE",
        )
        warning_fill = self.fill_green if missing_price_count == 0 else self.fill_red
        for row_idx in (total_row + 1, total_row + 2):
            for col_idx in (label_col_idx, warning_col_idx):
                self.ws.cell(row=row_idx, column=col_idx).font = header_font
                self.ws.cell(row=row_idx, column=col_idx).fill = warning_fill

        # Enable AutoFilter for the header row
        self.ws.auto_filter.ref = self.ws.dimensions
        # Freeze the top header row
        self.ws.freeze_panes = "A2"

        self._add_summary_sheet()
        self._add_cost_sheets()
        self._add_supplier_stock_sheet()
        add_refresh_changes_sheet(self.wb, self.items)
        atomic_save_workbook(self.wb, output_path)

    @staticmethod
    def _is_price_missing(item: BomItem) -> bool:
        status_lower = str(item.status or "").lower()
        jlc_usable = (
            bool(item.jlcpcb_part_number)
            and item.unit_price is not None
            and not item.is_jlcpcb_preorder
            and not any(
                marker in status_lower
                for marker in ("not found", "no exact", "error")
            )
        )
        return not jlc_usable and item.digikey_unit_price is None

    @staticmethod
    def _selected_supplier_price(
        item: BomItem,
        jlcpcb_price: Optional[float],
        digikey_price: Optional[float],
        required_quantity: int,
    ) -> tuple[str, Optional[float]]:
        """Return a stock-sufficient supplier for legacy Excel callers."""
        if (
            item.jlcpcb_part_number
            and not item.is_jlcpcb_preorder
            and jlcpcb_price is not None
            and item.available_stock_qty is not None
            and item.available_stock_qty >= required_quantity
        ):
            return "JLCPCB", jlcpcb_price
        if (
            item.digikey_part_number
            and digikey_price is not None
            and item.digikey_stock_qty is not None
            and item.digikey_stock_qty >= required_quantity
        ):
            return "DigiKey", digikey_price
        return "None", None

    def _add_supplier_stock_sheet(self):
        ws = self.wb.create_sheet("Supplier Stock")
        headers = [
            "MPN", "Design Item ID", "Supplier", "Supplier Part Number",
            "Stock", "Unit Price", "Production Required Quantity", "Safety Surplus",
            "Purchase Quantity", "Total Price", "Required Stock", "Status",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for item in self.items:
            prod_qty = item.production_quantity_val
            surplus = item.safety_surplus
            purchase_qty = item.purchase_quantity_val
            rows = [
                (
                    "JLCPCB", item.jlcpcb_part_number, item.available_stock_qty,
                    None if item.is_jlcpcb_preorder else item.unit_price,
                    None if item.is_jlcpcb_preorder else item.jlcpcb_total_price,
                    "Pre-order" if item.is_jlcpcb_preorder else item.status,
                ),
                ("DigiKey", item.digikey_part_number, item.digikey_stock_qty, item.digikey_unit_price, item.digikey_total_price,
                 "Found" if item.digikey_part_number else "Not Found"),
            ]
            for supplier, part_number, stock, unit_price, total_price, status in rows:
                ws.append([
                    item.mpn, item.comment, supplier, part_number or "-",
                    stock if stock is not None else "-", unit_price,
                    prod_qty, surplus, purchase_qty, total_price,
                    item.required_stock, status or "",
                ])
                if isinstance(unit_price, (int, float)):
                    ws.cell(row=ws.max_row, column=6).number_format = '"$"#,##0.0000'
                if isinstance(total_price, (int, float)):
                    ws.cell(row=ws.max_row, column=10).number_format = '"$"#,##0.0000'
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for column, width in {"A": 24, "B": 20, "C": 12, "D": 26, "E": 14, "F": 16, "G": 16, "H": 28}.items():
            ws.column_dimensions[column].width = width

    def _add_summary_sheet(self):
        """Add a secondary sheet with high-level sourcing statistics."""
        ws_sum = self.wb.create_sheet("Sourcing Summary")
        
        ws_sum.append(["Metric", "Count"])
        ws_sum["A1"].font = Font(bold=True)
        ws_sum["B1"].font = Font(bold=True)

        total = len(self.items)
        valid = sum(1 for i in self.items if i.status == "" and i.jlcpcb_part_number)
        yellow = sum(1 for i in self.items if i.status != "" and "error" not in i.status.lower())
        red = sum(1 for i in self.items if "error" in i.status.lower())

        ws_sum.append(["Total Components", total])
        ws_sum.append(["Valid (Green)", valid])
        ws_sum.append(["Manual Review (Yellow)", yellow])
        ws_sum.append(["API Errors (Red)", red])
        missing_price_count = sum(1 for item in self.items if self._is_price_missing(item))
        ws_sum.append(["Missing Price Count", missing_price_count])
        ws_sum.append([
            "Cost Status",
            "COMPLETE" if missing_price_count == 0 else "INCOMPLETE",
        ])
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = self.fill_green if missing_price_count == 0 else self.fill_red

        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 15

    def _add_cost_sheets(self):
        import re
        groups = {}
        for item in self.items:
            src = item.source_file_name if item.source_file_name else item.board_name
            if not src:
                src = "Unknown"
            groups.setdefault(src, []).append(item)
            
        used_names = {"Enriched BOM", "Sourcing Summary"}
        
        for src, group_items in groups.items():
            base_name = re.sub(r'[\\\\/?*\\[\\]:]', '', src)
            base_name = base_name.replace(".xlsx", "").replace(".csv", "").strip()
            base_name = f"{base_name} Cost"
            base_name = base_name[:31]
            
            sheet_name = base_name
            counter = 2
            while sheet_name in used_names:
                suffix = f"_{counter}"
                sheet_name = base_name[:31 - len(suffix)] + suffix
                counter += 1
            
            used_names.add(sheet_name)
            self._write_single_cost_sheet(sheet_name, group_items)

    def _write_single_cost_sheet(self, sheet_name: str, items: List[BomItem]):
        from core.mpn_utils import select_unit_price, select_digikey_price
        ws = self.wb.create_sheet(sheet_name)
        
        # --- Summary Table ---
        sum_headers = [
            "BOM Quantity", "JLCPCB Total", "DigiKey Fallback Total",
            "Mixed Sourcing Total", "All-DigiKey Total", "Missing Price Count",
            "Cost Status",
        ]
        ws.append(sum_headers)
        
        multipliers = self.build_multipliers
        summary_rows = {}
        for m in multipliers:
            summary_rows[m] = {
                "jlc_total": 0.0,
                "rem_dk_total": 0.0,
                "all_dk_total": 0.0,
                "missing_count": 0
            }
            
        header_font = Font(bold=True)
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            
        # Pre-calculate data for each row
        row_data_cache = []
        for item in items:
            row_cache = {}
            for m in multipliers:
                try:
                    scaled_qty = parse_positive_integer_quantity(item.quantity) * m
                except ValueError:
                    scaled_qty = 0
                surplus = item.safety_surplus
                purchase_qty = scaled_qty + surplus if scaled_qty > 0 else 0
                
                # Fetch JLC and DK prices
                j_price = None
                if item.jlcpcb_part_number and not item.is_jlcpcb_preorder and "error" not in item.status.lower() and "not found" not in item.status.lower() and "mismatch" not in item.status.lower():
                    j_price = select_unit_price(item.jlcpcb_price_breaks_raw, purchase_qty, use_quantity_breaks=self.pricing_mode == "project")
                    if j_price is None:
                        j_price = item.unit_price
                    
                d_price = select_digikey_price(item.digikey_price_breaks, purchase_qty, use_quantity_breaks=self.pricing_mode == "project")
                if d_price is None:
                    d_price = item.digikey_unit_price
                
                used_source = "Missing price"
                line_total = None
                
                j_purchasable = (
                    bool(item.jlcpcb_part_number)
                    and not item.is_jlcpcb_preorder
                    and j_price is not None
                    and "error" not in item.status.lower()
                    and "not found" not in item.status.lower()
                    and "mismatch" not in item.status.lower()
                    and (item.available_stock_qty is None or item.available_stock_qty >= purchase_qty)
                    and purchase_qty > 0
                )
                d_purchasable = (
                    bool(item.digikey_part_number)
                    and d_price is not None
                    and item.digikey_status != "error"
                    and (item.digikey_stock_qty is None or item.digikey_stock_qty >= purchase_qty)
                    and purchase_qty > 0
                )

                if j_purchasable and j_price is not None:
                    used_source = "JLCPCB"
                    line_total = purchase_qty * j_price
                    summary_rows[m]["jlc_total"] += line_total
                elif d_purchasable and d_price is not None:
                    used_source = "DigiKey fallback"
                    line_total = purchase_qty * d_price
                    summary_rows[m]["rem_dk_total"] += line_total
                else:
                    summary_rows[m]["missing_count"] += 1
                    
                if d_price is not None:
                    summary_rows[m]["all_dk_total"] += purchase_qty * d_price
                    
                row_cache[m] = {
                    "j_price": j_price,
                    "d_price": d_price,
                    "used_source": used_source,
                    "line_total": line_total
                }
            row_data_cache.append(row_cache)
            
        # Write Summary Data
        for idx, m in enumerate(multipliers, 2):
            jlc_t = summary_rows[m]["jlc_total"]
            rem_dk_t = summary_rows[m]["rem_dk_total"]
            comb_t = jlc_t + rem_dk_t
            all_dk_t = summary_rows[m]["all_dk_total"]
            miss_c = summary_rows[m]["missing_count"]
            
            row_vals = [
                f"{m}x", jlc_t, rem_dk_t, comb_t, all_dk_t, miss_c,
                "COMPLETE" if miss_c == 0 else "INCOMPLETE",
            ]
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=idx, column=c_idx, value=val)
                if c_idx in (2, 3, 4, 5):
                    cell.number_format = "#,##0.0000"
                if c_idx in (6, 7):
                    cell.fill = self.fill_green if miss_c == 0 else self.fill_red
                    
        # --- Detailed Table ---
        det_start_row = len(multipliers) + 3
        
        det_headers = [
            "Designator", "Quantity", "Value", "Manufacturer", "MPN",
            "JLCPCB Part Number", "Status"
        ]
        for m in multipliers:
            det_headers.extend([
                f"{m}x JLCPCB Unit",
                f"{m}x DigiKey Unit",
                f"{m}x Used Source",
                f"{m}x Line Total"
            ])
            
        for c_idx, h_text in enumerate(det_headers, 1):
            cell = ws.cell(row=det_start_row, column=c_idx, value=h_text)
            cell.font = header_font
            
        for r_idx, (item, cache) in enumerate(zip(items, row_data_cache), det_start_row + 1):
            row_vals = [
                item.designator,
                item.quantity,
                item.value,
                item.manufacturer,
                item.mpn,
                item.jlcpcb_part_number,
                item.status if item.status else "Valid"
            ]
            
            for m in multipliers:
                rc = cache[m]
                row_vals.extend([
                    rc["j_price"] if rc["j_price"] is not None else "-",
                    rc["d_price"] if rc["d_price"] is not None else "-",
                    rc["used_source"],
                    rc["line_total"] if rc["line_total"] is not None else "-"
                ])
                
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # Apply format
                if c_idx == 2:
                    cell.number_format = "0"
                elif c_idx > 7:
                    rel_idx = (c_idx - 8) % 4
                    if rel_idx in (0, 1, 3) and isinstance(val, (int, float)):
                        cell.number_format = "#,##0.0000"
                        
        # Formatting
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 25
        
        ws.auto_filter.ref = f"A{det_start_row}:{get_column_letter(len(det_headers))}{det_start_row + len(items)}"
        ws.freeze_panes = f"A{det_start_row + 1}"

class UnavailableReportWriter:
    """Write items that have no usable catalogue code from either supplier."""

    def __init__(self, items: List[BomItem]):
        # JLCPCB status describes only one supplier. A DigiKey exact match still
        # makes the component sourceable even when JLCPCB reports "not found".
        self.unavailable_items = [
            item
            for item in items
            if (not item.jlcpcb_part_number or item.is_jlcpcb_preorder)
            and not item.digikey_part_number
        ]

    def write(self, output_path: str) -> None:
        """Write the unavailable report data to the given Excel file path."""
        if not self.unavailable_items:
            return  # Nothing to write

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Action Required"

        headers = [
            "Board Name", "Designator", "Quantity", "Manufacturer", 
            "MPN", "Required Stock", "JLCPCB Stock", "Status"
        ]
        ws.append(headers)

        header_font = Font(bold=True)
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        ws.column_dimensions["B"].width = 30 # Designator
        ws.column_dimensions["H"].width = 30 # Status

        for r_idx, item in enumerate(self.unavailable_items, 2):
            row = [
                item.board_name,
                item.designator,
                item.quantity,
                item.manufacturer,
                item.mpn,
                item.required_stock,
                item.available_stock_qty if item.available_stock_qty is not None else "-",
                item.status
            ]
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Enable AutoFilter and freeze panes
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        atomic_save_workbook(wb, output_path)

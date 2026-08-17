import os
import pytest
import openpyxl

from models.bom_item import BomItem
from models.project import Project, ProjectItem
from models.workspace import Workspace
from services.project_aggregation import aggregate_workspace
from core.workspace_excel_writer import WorkspaceExcelWriter

def test_workbook_structure_and_sheet_names(tmp_path):
    workspace = Workspace("Test WS")
    
    # Project 1
    p1 = Project("P1")
    p1.add_board(ProjectItem("path_a/board.xlsx", "Board A", 1))
    workspace.add_project(p1)
    
    # Project 2 (Same board name to test collision)
    p2 = Project("P2")
    p2.add_board(ProjectItem("path_b/board.xlsx", "Board A", 2))
    workspace.add_project(p2)
    
    agg = aggregate_workspace(workspace)
    writer = WorkspaceExcelWriter(workspace, agg, [], [])
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    names = wb.sheetnames
    
    # Required Sheets
    assert "All Aggregated" in names
    assert "Mutual Components" in names
    assert names[0] == "All Aggregated"
    assert names[1:3] == ["Project Summary - P1", "Project Summary - P2"]
    assert wb.active.title == "All Aggregated"
    
    # Projects
    assert "Project Summary - P1" in names
    assert "Project Summary - P2" in names
    
    # Name constraints validation
    for name in names:
        assert len(name) <= 31
        for invalid_char in [':', '\\', '/', '?', '*', '[', ']']:
            assert invalid_char not in name

def test_strict_validation(tmp_path):
    workspace = Workspace("Test")
    p1 = Project("P1")
    b = ProjectItem("path", "B", 1)
    b.bom_items = [BomItem(mpn="R1", quantity=1)]
    p1.add_board(b)
    workspace.add_project(p1)
    
    agg = aggregate_workspace(workspace)
    c_keys = [c.component_key for c in agg.components]
    
    enriched = [BomItem(mpn="R1")]
    
    # Success
    WorkspaceExcelWriter(workspace, agg, enriched, c_keys)
    
    # Length mismatch
    with pytest.raises(ValueError, match="Length mismatch"):
        WorkspaceExcelWriter(workspace, agg, enriched, c_keys + ["extra"])
        
    # Duplicates in component keys
    with pytest.raises(ValueError, match="Duplicate component keys"):
        WorkspaceExcelWriter(workspace, agg, enriched + enriched, c_keys + c_keys)
        
    # Missing / Extra (Mismatched from aggregation)
    with pytest.raises(ValueError, match="Component key mismatch"):
        WorkspaceExcelWriter(workspace, agg, enriched, ["WRONG_KEY"])


def test_mutual_components_sheet(tmp_path):
    workspace = Workspace("Test")
    
    p1 = Project("P1")
    b1 = ProjectItem("path1", "B1", 1)
    b1.bom_items = [
        BomItem(mpn="MUTUAL", quantity=2), 
        BomItem(mpn="P1_ONLY", quantity=1)
    ]
    p1.add_board(b1)
    workspace.add_project(p1)
    
    p2 = Project("P2")
    b2 = ProjectItem("path2", "B2", 3)
    b2.bom_items = [
        BomItem(mpn="MUTUAL", quantity=1)
    ]
    p2.add_board(b2)
    workspace.add_project(p2)
    
    agg = aggregate_workspace(workspace)
    c_keys = [c.component_key for c in agg.components]
    enriched = []
    for k in c_keys:
        if "MUTUAL" in k:
            enriched.append(BomItem(mpn="MUTUAL", manufacturer="MFG"))
        else:
            enriched.append(BomItem(mpn="P1_ONLY"))
            
    writer = WorkspaceExcelWriter(workspace, agg, enriched, c_keys, build_multipliers=[1])
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    assert "Mutual Components" in wb.sheetnames
    
    ws = wb["Mutual Components"]
    headers = [cell.value for cell in ws[1]]
    mpn_col_idx = headers.index("MPN")
    rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row[mpn_col_idx]
    ]
    
    # MUTUAL now merged into 1 row
    assert len(rows) == 1
    
    board_col_idx = headers.index("Board Names")
    qty_col_idx = headers.index("Quantity (Per Board / Total)")
    assert rows[0][mpn_col_idx] == "MUTUAL"
    
    assert rows[0][board_col_idx] == "B1\nB2"
    assert rows[0][qty_col_idx] == "2 / 2\n1 / 3"


def test_workspace_writes_each_actual_board_to_a_separate_sheet(tmp_path):
    workspace = Workspace("Separated WS")
    project = Project("Device")

    combined_file = ProjectItem("files/combined.xlsx", "combined", 1)
    combined_file.bom_items = [
        BomItem(board_name="battery", mpn="A", quantity=1, designator="A1"),
        BomItem(board_name="battery", mpn="B", quantity=2, designator="B1"),
        BomItem(board_name="main", mpn="B", quantity=3, designator="B2"),
        BomItem(board_name="main", mpn="G", quantity=1, designator="G1"),
    ]
    project.add_board(combined_file)
    workspace.add_project(project)

    agg = aggregate_workspace(workspace)
    component_keys = [component.component_key for component in agg.components]
    enriched = [
        BomItem(mpn=component.representative_item.mpn)
        for component in agg.components
    ]

    output_path = tmp_path / "separated.xlsx"
    WorkspaceExcelWriter(
        workspace,
        agg,
        enriched,
        component_keys,
        build_multipliers=[1],
    ).write(str(output_path))

    workbook = openpyxl.load_workbook(output_path, data_only=False)
    assert workbook.sheetnames[:4] == [
        "All Aggregated",
        "Project Summary - Device",
        "battery",
        "main",
    ]

    def sheet_mpns(sheet_name):
        sheet = workbook[sheet_name]
        header_row = next(
            row[0].row for row in sheet.iter_rows()
            if row[0].value == "Board Name"
        )
        headers = [cell.value for cell in sheet[header_row]]
        mpn_index = headers.index("MPN")
        return {
            row[mpn_index]
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True)
            if row[mpn_index]
        }

    assert sheet_mpns("battery") == {"A", "B"}
    assert sheet_mpns("main") == {"B", "G"}

    battery_sheet = workbook["battery"]
    assert battery_sheet["P11"].value == "Board Total"
    assert battery_sheet["Q11"].value == "=SUM(Q8:Q9)"
    assert battery_sheet["R11"].value == "=SUM(R8:R9)"
    assert battery_sheet["P12"].value == "Cost Per Board Set"
    assert battery_sheet["Q12"].value == "=Q11/1"
    assert battery_sheet["R12"].value == "=R11/1"


def test_board_price_tier_uses_total_quantity_across_project_boards(tmp_path):
    workspace = Workspace("Pooled Pricing")
    project = Project("Device")
    combined = ProjectItem("combined.xlsx", "combined", 1)
    combined.bom_items = [
        BomItem(board_name="Board A", mpn="SHARED", quantity=4),
        BomItem(board_name="Board B", mpn="SHARED", quantity=3),
        BomItem(board_name="Board C", mpn="SHARED", quantity=3),
    ]
    project.add_board(combined)
    workspace.add_project(project)

    agg = aggregate_workspace(workspace)
    component = agg.components[0]
    enriched = BomItem(
        mpn="SHARED",
        jlcpcb_price_breaks_raw='[{"qFrom": 1, "price": 5.0}, {"qFrom": 10, "price": 3.5}]',
        digikey_price_breaks=[(1, 5.0), (10, 3.5)],
    )
    output_path = tmp_path / "pooled.xlsx"
    WorkspaceExcelWriter(
        workspace,
        agg,
        [enriched],
        [component.component_key],
        build_multipliers=[1],
        pricing_mode="project",
    ).write(str(output_path))

    workbook = openpyxl.load_workbook(output_path, data_only=False)
    for board_name, required_quantity in (("Board A", 4), ("Board B", 3), ("Board C", 3)):
        sheet = workbook[board_name]
        headers = [cell.value for cell in sheet[7]]
        row = [cell.value for cell in sheet[8]]
        assert row[headers.index("Board Required Quantity")] == required_quantity
        assert row[headers.index("Pricing Pool Quantity")] == 10
        assert row[headers.index("JLCPCB Unit Price")] == 3.5
        assert row[headers.index("DigiKey Unit Price")] == 3.5
        assert row[headers.index("JLCPCB Total Price")] == required_quantity * 3.5


def test_named_board_sheet_contains_its_own_supplier_prices(tmp_path):
    workspace = Workspace("Satellite")
    project = Project("SAT6")
    board_name = "PBA_SAT6_CB_UWB_P0_B0"
    board = ProjectItem("boards/uwb.xlsx", board_name, 2)
    board.bom_items = [BomItem(board_name=board_name, mpn="UWB-MPN", quantity=3)]
    project.add_board(board)
    workspace.add_project(project)

    agg = aggregate_workspace(workspace)
    component = agg.components[0]
    enriched = BomItem(
        mpn="UWB-MPN",
        jlcpcb_part_number="C123",
        digikey_part_number="DK-123",
        jlcpcb_price_breaks_raw='[{"qFrom": 1, "price": 2.0}]',
        digikey_price_breaks=[(1, 2.5)],
    )
    output_path = tmp_path / "named-board.xlsx"
    WorkspaceExcelWriter(
        workspace,
        agg,
        [enriched],
        [component.component_key],
        build_multipliers=[5],
    ).write(str(output_path))

    workbook = openpyxl.load_workbook(output_path, data_only=False)
    sheet = workbook[board_name]
    headers = [cell.value for cell in sheet[7]]
    row = [cell.value for cell in sheet[8]]

    assert sheet["B4"].value == 10
    assert row[headers.index("Board Required Quantity")] == 30
    assert row[headers.index("JLCPCB Unit Price")] == 2.0
    assert row[headers.index("DigiKey Unit Price")] == 2.5
    assert row[headers.index("JLCPCB Total Price")] == 60.0
    assert row[headers.index("DigiKey Total Price")] == 75.0
    assert row[headers.index("Selected Supplier")] == "JLCPCB"
    assert row[headers.index("Mixed Sourcing Total Price")] == 60.0

    total_row = 10
    per_board_set_row = 11
    mixed_total_col = headers.index("Mixed Sourcing Total Price") + 1
    assert sheet.cell(total_row, mixed_total_col).value == f"=SUM(T8:T8)"
    assert sheet.cell(per_board_set_row, mixed_total_col).value == f"=T10/5"
    assert sheet["D1"].value == "Card Production Cost (Mixed Sourcing):"
    assert sheet["E1"].value == "=SUM(T8:T8)"
    assert sheet["D2"].value == "Cost Per Card:"
    assert sheet["E2"].value == "=E1/10"


def test_board_production_quantity_is_build_quantity_when_bom_file_quantity_is_one(tmp_path):
    workspace = Workspace("Single Board")
    project = Project("Device")
    board = ProjectItem("boards/main.xlsx", "PBA_SAR_CB", 1)
    board.bom_items = [BomItem(board_name="PBA_SAR_CB", mpn="PART", quantity=1)]
    project.add_board(board)
    workspace.add_project(project)

    agg = aggregate_workspace(workspace)
    component = agg.components[0]
    output_path = tmp_path / "single-board.xlsx"
    WorkspaceExcelWriter(
        workspace,
        agg,
        [BomItem(mpn="PART")],
        [component.component_key],
        build_multipliers=[5],
    ).write(str(output_path))

    workbook = openpyxl.load_workbook(output_path, data_only=False)
    sheet = workbook["PBA_SAR_CB"]
    headers = [cell.value for cell in sheet[7]]
    row = [cell.value for cell in sheet[8]]
    assert sheet["B4"].value == 5
    assert row[headers.index("Board Required Quantity")] == 5



def test_pricing_fallback_and_unpriced(tmp_path):
    workspace = Workspace("Test")
    p1 = Project("P1")
    b = ProjectItem("path", "B", 1)
    b.bom_items = [
        BomItem(mpn="UNPRICED", quantity=1),
        BomItem(mpn="FALLBACK", quantity=1),
    ]
    p1.add_board(b)
    workspace.add_project(p1)
    
    agg = aggregate_workspace(workspace)
    c_keys = [c.component_key for c in agg.components]
    
    enriched = []
    for k in c_keys:
        if "UNPRICED" in k:
            enriched.append(BomItem(mpn="UNPRICED", status="Not found"))
        else:
            enriched.append(BomItem(
                mpn="FALLBACK", 
                jlcpcb_part_number="C1",
                status="Insufficient JLCPCB stock",
                jlcpcb_price_breaks_raw='[{"qFrom": 1, "price": 0.10}]',
                digikey_price_breaks=[(1, 0.50)]
            ))
            
    writer = WorkspaceExcelWriter(workspace, agg, enriched, c_keys, build_multipliers=[1])
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    ws = wb["All Aggregated"]

    board_ws = wb["B"]
    assert board_ws["E3"].value == 1
    assert board_ws["E4"].value == "INCOMPLETE"

    project_ws = wb["Project Summary - P1"]
    project_cost_row = next(
        row for row in project_ws.iter_rows(values_only=True) if row[0] == "1x"
    )
    assert project_cost_row[5] == 1
    assert project_cost_row[6] == "INCOMPLETE"
    
    # Headers
    headers = [cell.value for cell in ws[1]]
    
    # Regression test for absent bloated headers
    assert "Selected Source @ 5x" not in headers
    assert "Selected Source @ 1x" not in headers
    assert "JLCPCB Extended Cost @ 5x" not in headers
    
    mpn_col_idx = headers.index("MPN")
    jlc_part_col_idx = headers.index("JLCPCB Part Number")
    status_col_idx = headers.index("Status")
    
    found_unpriced = False
    found_fallback = False
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        mpn = row[mpn_col_idx]
        if mpn == "UNPRICED":
            assert row[status_col_idx] == "Not found"
            found_unpriced = True
        elif mpn == "FALLBACK":
            assert row[jlc_part_col_idx] == "C1"
            assert row[status_col_idx] == "Insufficient JLCPCB stock"
            found_fallback = True
            
    assert found_unpriced
    assert found_fallback

def test_workspace_multiboard_quantity_formatting_and_auto_width(tmp_path):
    long_desc = "Precision resistor network for controller power measurement"
    workspace = Workspace("Test")
    
    project = Project("P1")
    board_a = ProjectItem("path_a", "Board A", 2)
    board_a.bom_items = [
        BomItem(mpn="R1", quantity=2.0, description=long_desc, designator="R1")
    ]
    board_b = ProjectItem("path_b", "Board B", 1)
    board_b.bom_items = [
        BomItem(mpn="R1", quantity="3.0", description=long_desc, designator="R2")
    ]
    project.add_board(board_a)
    project.add_board(board_b)
    workspace.add_project(project)
    
    agg = aggregate_workspace(workspace)
    c_keys = [c.component_key for c in agg.components]
    enriched = [
        BomItem(mpn="R1", description=long_desc)
    ]
    
    writer = WorkspaceExcelWriter(workspace, agg, enriched, c_keys, build_multipliers=[1])
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    all_ws = wb["All Aggregated"]
    assert all_ws["A1"].value == "Board Name"
    assert all_ws["A2"].value == "Board A"
    assert all_ws["C2"].value == "R1"
    assert all_ws["D2"].value == "2 / 4"
    assert all_ws["A3"].value == "Board B"
    assert all_ws["C3"].value == "R2"
    assert all_ws["D3"].value == "3 / 3"
    assert all_ws.column_dimensions["B"].width == 35
    
    project_ws = wb["Project Summary - P1"]
    project_rows = list(project_ws.iter_rows(values_only=True))
    project_component_row_a = next(row for row in project_rows if row[0] == "Board A")
    assert project_component_row_a[2] == "R1"
    assert project_component_row_a[3] == "2 / 4"
    
    project_component_row_b = next(row for row in project_rows if row[0] == "Board B")
    assert project_component_row_b[2] == "R2"
    assert project_component_row_b[3] == "3 / 3"
    

import pytest
from openpyxl import Workbook
from PyQt6.QtWidgets import QApplication

from core.database_manager import DatabaseManager
from core.component_library import (
    read_component_library_file,
    detect_library_conflicts,
    ConflictItem,
)
from ui.component_library_conflict_dialog import ComponentLibraryConflictDialog


@pytest.fixture(scope="session")
def qapp():
    app = QApplication.instance()
    if app is None:
        app = QApplication([])
    return app


def test_identical_mpn_normalization_merges_safely(tmp_path):
    path = tmp_path / "merge_identical.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Components"
    ws.append(["LIBRARYREFERENCE", "MANUFACTURER PART NUMBER"])
    ws.append(["RES010001", "RC0603FR-0710KL"])
    ws.append(["RES010001", "  rc0603fr-0710kl \n"])  # Same MPN with different case/whitespace
    ws.append(["CAP010001", "GRM188R71C104KA01D"])
    wb.save(path)

    raw_rows, invalid_skipped = read_component_library_file(str(path))
    assert len(raw_rows) == 3
    assert invalid_skipped == 0

    clean, conflicts, merged = detect_library_conflicts(raw_rows)
    assert len(conflicts) == 0
    assert merged == 1  # 1 duplicate occurrence merged
    assert len(clean) == 2
    assert ("RES010001", "RC0603FR-0710KL") in clean
    assert ("CAP010001", "GRM188R71C104KA01D") in clean


def test_different_mpns_for_same_internal_code_are_flagged_as_file_conflict(tmp_path):
    path = tmp_path / "conflict_file.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["LIBRARYREFERENCE", "MANUFACTURER PART NUMBER"])
    ws.append(["RES010001", "RC0603FR-0710KL"])
    ws.append(["RES010001", "ERJ-3EKF1002V"])  # Different MPN for same code
    ws.append(["IC010001", "STM32F401RE"])
    wb.save(path)

    raw_rows, _ = read_component_library_file(str(path))
    clean, conflicts, _ = detect_library_conflicts(raw_rows)

    # Clean components must only contain the non-conflicting IC010001
    assert clean == [("IC010001", "STM32F401RE")]

    # RES010001 must be in conflicts
    assert len(conflicts) == 1
    conflict = conflicts[0]
    assert conflict.internal_code == "RES010001"
    assert conflict.conflict_type == "FILE_CONFLICT"
    assert set(conflict.candidate_mpns) == {"RC0603FR-0710KL", "ERJ-3EKF1002V"}


def test_file_mpn_differing_from_db_is_flagged_as_db_conflict(tmp_path):
    db_path = str(tmp_path / "db.sqlite")
    db = DatabaseManager(db_path)
    # Seed existing mapping in database
    db.upsert_internal_mapping("RES010001", "RC0603FR-0710KL", "C123", True)

    path = tmp_path / "conflict_db.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["COMMENT", "MANUFACTURER PART NUMBER"])
    ws.append(["RES010001", "ERJ-3EKF1002V"])  # Different from DB's RC0603FR-0710KL
    ws.append(["CAP010001", "GRM188R71C104KA01D"])
    wb.save(path)

    raw_rows, _ = read_component_library_file(str(path))
    clean, conflicts, _ = detect_library_conflicts(raw_rows, db)

    assert clean == [("CAP010001", "GRM188R71C104KA01D")]
    assert len(conflicts) == 1
    conflict = conflicts[0]
    assert conflict.internal_code == "RES010001"
    assert conflict.conflict_type == "DB_CONFLICT"
    assert conflict.existing_db_mpn == "RC0603FR-0710KL"
    assert conflict.candidate_mpns == ["ERJ-3EKF1002V"]


def test_same_mpn_as_db_is_not_flagged_as_conflict(tmp_path):
    db_path = str(tmp_path / "db.sqlite")
    db = DatabaseManager(db_path)
    db.upsert_internal_mapping("RES010001", "RC0603FR-0710KL", "C123", True)

    path = tmp_path / "match_db.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["LIBRARYREFERENCE", "MANUFACTURER PART NUMBER"])
    ws.append(["RES010001", "  rc0603fr-0710kl  "])  # Same MPN after normalization
    wb.save(path)

    raw_rows, _ = read_component_library_file(str(path))
    clean, conflicts, _ = detect_library_conflicts(raw_rows, db)

    assert len(conflicts) == 0
    assert len(clean) == 1
    assert clean[0][0] == "RES010001"


def test_conflict_dialog_defaults_to_skip_and_resolves_chosen(qapp):
    conflict1 = ConflictItem(
        internal_code="RES010001",
        conflict_type="FILE_CONFLICT",
        candidate_mpns=["RC0603FR-0710KL", "ERJ-3EKF1002V"],
        row_numbers={"RC0603FR-0710KL": [2], "ERJ-3EKF1002V": [3]},
    )
    conflict2 = ConflictItem(
        internal_code="CAP010001",
        conflict_type="DB_CONFLICT",
        candidate_mpns=["GRM-NEW"],
        row_numbers={"GRM-NEW": [4]},
        existing_db_mpn="GRM-OLD",
    )

    dlg = ComponentLibraryConflictDialog([conflict1, conflict2])

    # Default state: both combos on index 0 (Skip)
    assert len(dlg.get_resolved_mappings()) == 0

    # User explicitly chooses candidate 1 ("RC0603FR-0710KL") for RES010001
    combo1 = dlg._combos["RES010001"]
    combo1.setCurrentIndex(1)

    # CAP010001 is left on Skip
    resolved = dlg.get_resolved_mappings()
    assert len(resolved) == 1
    assert resolved[0] == ("RES010001", "RC0603FR-0710KL")

    # "Use First File MPN for All"
    dlg._use_first_for_all()
    resolved_all = dlg.get_resolved_mappings()
    assert len(resolved_all) == 2
    assert resolved_all[0] == ("RES010001", "RC0603FR-0710KL")
    assert resolved_all[1] == ("CAP010001", "GRM-NEW")

    # "Skip All"
    dlg._skip_all()
    assert len(dlg.get_resolved_mappings()) == 0


def test_db_layer_duplicate_code_deduplication(tmp_path):
    db_path = str(tmp_path / "db.sqlite")
    db = DatabaseManager(db_path)

    # Pass records with duplicate comment_codes
    records = [
        ("RES01", "MPN-1", "C1", "DK1"),
        ("RES01", "MPN-2", "C2", "DK2"),  # Duplicate comment_code
        ("CAP01", "CAP-1", "C3", "DK3"),
    ]

    inserted = db.bulk_insert_new_pending_suggestions(records)
    assert inserted == 2

    # Verify only first RES01 was inserted
    mapping = db.get_internal_mapping("RES01")
    assert mapping is not None
    assert mapping["mpn"] == "MPN-1"


def test_db_layer_insert_pending_preserves_existing_mpn(tmp_path):
    db_path = str(tmp_path / "db.sqlite")
    db = DatabaseManager(db_path)

    # Seed record with existing MPN
    db.upsert_internal_mapping("RES01", "ORIGINAL-MPN", "C1", True)

    # Attempt to insert pending suggestion with different MPN
    db.insert_pending_suggestion("RES01", "DIFFERENT-MPN", "C2", "DK2")

    # Verify original MPN was not overwritten
    mapping = db.get_internal_mapping("RES01")
    assert mapping is not None
    assert mapping["mpn"] == "ORIGINAL-MPN"

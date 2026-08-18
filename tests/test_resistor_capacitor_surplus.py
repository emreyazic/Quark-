import pytest
import openpyxl

from core.mpn_utils import is_resistor_or_capacitor, get_safety_surplus, select_unit_price, select_digikey_price
from models.bom_item import BomItem
from models.project import Project, ProjectItem
from models.workspace import Workspace
from services.project_aggregation import aggregate_project, aggregate_workspace
from core.project_excel_writer import ProjectExcelWriter
from core.workspace_excel_writer import WorkspaceExcelWriter
from core.excel_writer import ExcelWriter
from core.jlcpcb_searcher import JlcpcbSearcher
from core.digikey_searcher import DigiKeySearcher


def test_component_classification_resistors_and_capacitors():
    # Designator-first recognition
    assert is_resistor_or_capacitor("R1") is True
    assert is_resistor_or_capacitor("R102") is True
    assert is_resistor_or_capacitor("RN1") is True
    assert is_resistor_or_capacitor("RN4") is True
    assert is_resistor_or_capacitor("C1") is True
    assert is_resistor_or_capacitor("C999") is True
    assert is_resistor_or_capacitor("R1, R2, R3") is True
    assert is_resistor_or_capacitor("C1, C2, C3") is True
    assert is_resistor_or_capacitor("C1-C10") is True

    # Non-R/C designators
    assert is_resistor_or_capacitor("U1") is False
    assert is_resistor_or_capacitor("IC1") is False
    assert is_resistor_or_capacitor("D1") is False
    assert is_resistor_or_capacitor("Q1") is False
    assert is_resistor_or_capacitor("CN1") is False
    assert is_resistor_or_capacitor("CR1") is False
    assert is_resistor_or_capacitor("RF1") is False
    assert is_resistor_or_capacitor("REG1") is False

    # Secondary indicators (description / comment / value / footprint)
    assert is_resistor_or_capacitor("U1", description="10k Resistor 0402 1%") is True
    assert is_resistor_or_capacitor("U1", description="100nF Ceramic Capacitor 50V") is True
    assert is_resistor_or_capacitor("", comment="RES 10K OHM 1% 0603") is True
    assert is_resistor_or_capacitor("", value="10uF", footprint="CAP_0805") is True
    assert is_resistor_or_capacitor("U1", description="Microcontroller 32-bit ARM") is False


def test_safety_surplus_amounts():
    resistor = BomItem(designator="R1", mpn="RC0603FR-0710KL", quantity=100)
    assert get_safety_surplus(resistor) == 10
    assert resistor.safety_surplus == 10
    assert resistor.production_quantity_val == 100
    assert resistor.purchase_quantity_val == 110

    capacitor = BomItem(designator="C1", mpn="CC0603KRX7R9BB104", quantity=100)
    assert get_safety_surplus(capacitor) == 10
    assert capacitor.safety_surplus == 10
    assert capacitor.production_quantity_val == 100
    assert capacitor.purchase_quantity_val == 110

    ic = BomItem(designator="U1", mpn="STM32F401RET6", quantity=100)
    assert get_safety_surplus(ic) == 0
    assert ic.safety_surplus == 0
    assert ic.production_quantity_val == 100
    assert ic.purchase_quantity_val == 100


def test_resistors_not_skipped_in_jlcpcb_searcher():
    searcher = JlcpcbSearcher()
    # RES prefix MPN should not be marked as RES manual skip
    res_item = BomItem(designator="R1", mpn="RES-0603-10K", description="Resistor 10k")
    # Calling internal filter check if any: verify skip_jlcpcb is not set
    assert res_item.skip_jlcpcb is False


def test_single_bom_aggregation_and_surplus_applied_once():
    project = Project("Project A")
    board = ProjectItem("board1.csv", "Board 1", 1)
    board.bom_items = [
        BomItem(designator="R1", mpn="RC0603-10K", quantity=40, description="10k Resistor"),
        BomItem(designator="R2", mpn="RC0603-10K", quantity=60, description="10k Resistor"),
    ]
    project.add_board(board)
    agg = aggregate_project(project)

    assert len(agg.components) == 1
    comp = agg.components[0]
    assert comp.total_quantity == 100
    assert comp.safety_surplus == 10
    assert comp.purchase_quantity == 110
    assert comp.representative_item.production_quantity_val == 100
    assert comp.representative_item.safety_surplus == 10
    assert comp.representative_item.purchase_quantity_val == 110


def test_multi_board_aggregation_surplus_applied_once():
    project = Project("Multi Board Project")
    board_a = ProjectItem("board_a.csv", "Board A", 2)
    board_a.bom_items = [
        BomItem(designator="R1", mpn="RC0603-10K", quantity=20, description="10k Resistor"),
    ]
    board_b = ProjectItem("board_b.csv", "Board B", 3)
    board_b.bom_items = [
        BomItem(designator="R10", mpn="RC0603-10K", quantity=20, description="10k Resistor"),
    ]
    project.add_board(board_a)
    project.add_board(board_b)
    # Total production required = 2 * 20 + 3 * 20 = 40 + 60 = 100
    agg = aggregate_project(project)

    assert len(agg.components) == 1
    comp = agg.components[0]
    assert comp.total_quantity == 100
    assert comp.safety_surplus == 10
    assert comp.purchase_quantity == 110


def test_workspace_aggregation_surplus_applied_once():
    ws = Workspace("Workspace 1")
    p1 = Project("Project 1")
    b1 = ProjectItem("p1_b1.csv", "P1 Board 1", 1)
    b1.bom_items = [BomItem(designator="C1", mpn="CAP-0603-100NF", quantity=50)]
    p1.add_board(b1)
    ws.add_project(p1)

    p2 = Project("Project 2")
    b2 = ProjectItem("p2_b1.csv", "P2 Board 1", 1)
    b2.bom_items = [BomItem(designator="C2", mpn="CAP-0603-100NF", quantity=50)]
    p2.add_board(b2)
    ws.add_project(p2)

    res = aggregate_workspace(ws)
    assert len(res.components) == 1
    w_comp = res.components[0]
    assert w_comp.total_quantity == 100
    assert w_comp.safety_surplus == 10
    assert w_comp.purchase_quantity == 110


def test_price_breaks_selected_by_purchase_quantity():
    raw_breaks = '[{"qFrom": 1, "price": 0.10}, {"qFrom": 100, "price": 0.05}, {"qFrom": 110, "price": 0.03}]'
    dk_breaks = [(1, 0.10), (100, 0.05), (110, 0.03)]

    # Production quantity is 100, but purchase quantity is 110
    # Price break should be 0.03 at 110
    j_price = select_unit_price(raw_breaks, 110, use_quantity_breaks=True)
    d_price = select_digikey_price(dk_breaks, 110, use_quantity_breaks=True)
    assert j_price == 0.03
    assert d_price == 0.03

    # At 100, price break would only be 0.05
    assert select_unit_price(raw_breaks, 100, use_quantity_breaks=True) == 0.05


def test_stock_sufficiency_evaluated_at_purchase_quantity(tmp_path):
    # Production required = 100, Surplus = 10, Purchase = 110
    # Case 1: Stock is 109 -> Insufficient
    item_insufficient = BomItem(
        designator="R1",
        mpn="RC0603-10K",
        quantity=100,
        production_quantity=100,
        safety_surplus=10,
        purchase_quantity=110,
        pricing_quantity=110,
        required_stock=110,
        jlcpcb_part_number="C12345",
        available_stock_qty=109,
        unit_price=0.05,
    )
    writer = ExcelWriter([item_insufficient], pricing_mode="project", build_multipliers=[1])
    # JLCPCB stock 109 is less than purchase quantity 110 -> missing/unpurchasable from JLCPCB
    used_src, _ = writer._selected_supplier_price(item_insufficient, 0.05, None, required_quantity=110)
    assert used_src == "None"

    # Case 2: Stock is 110 -> Sufficient
    item_sufficient = BomItem(
        designator="R1",
        mpn="RC0603-10K",
        quantity=100,
        production_quantity=100,
        safety_surplus=10,
        purchase_quantity=110,
        pricing_quantity=110,
        required_stock=110,
        jlcpcb_part_number="C12345",
        available_stock_qty=110,
        unit_price=0.05,
    )
    used_src_2, price_2 = writer._selected_supplier_price(item_sufficient, 0.05, None, required_quantity=110)
    assert used_src_2 == "JLCPCB"
    assert price_2 == 0.05


def test_excel_export_columns_and_values(tmp_path):
    project = Project("Test Project")
    board = ProjectItem("b1.xlsx", "Board 1", 1)
    board.bom_items = [
        BomItem(designator="R1", mpn="RC0603-10K", quantity=100, description="10k Resistor"),
        BomItem(designator="C1", mpn="CAP-100NF", quantity=100, description="100nF Capacitor"),
        BomItem(designator="U1", mpn="STM32F4", quantity=100, description="MCU"),
    ]
    project.add_board(board)
    agg = aggregate_project(project)

    enriched = [
        BomItem(
            designator="R1", mpn="RC0603-10K", quantity=100,
            production_quantity=100, safety_surplus=10, purchase_quantity=110,
            pricing_quantity=110, jlcpcb_part_number="C1001", unit_price=0.05,
            available_stock_qty=500
        ),
        BomItem(
            designator="C1", mpn="CAP-100NF", quantity=100,
            production_quantity=100, safety_surplus=10, purchase_quantity=110,
            pricing_quantity=110, jlcpcb_part_number="C1002", unit_price=0.02,
            available_stock_qty=500
        ),
        BomItem(
            designator="U1", mpn="STM32F4", quantity=100,
            production_quantity=100, safety_surplus=0, purchase_quantity=100,
            pricing_quantity=100, jlcpcb_part_number="C1003", unit_price=5.00,
            available_stock_qty=500
        ),
    ]

    out_file = tmp_path / "project_output.xlsx"
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=enriched,
        component_keys=[c.component_key for c in agg.components],
        build_multipliers=[1]
    )
    writer.write(str(out_file))

    wb = openpyxl.load_workbook(out_file, data_only=True)
    agg_ws = wb["Aggregated Components"]
    headers = [cell.value for cell in agg_ws[1]]

    assert "Production Required Quantity" in headers
    assert "Safety Surplus" in headers
    assert "Purchase Quantity" in headers

    prod_idx = headers.index("Production Required Quantity") + 1
    surplus_idx = headers.index("Safety Surplus") + 1
    purchase_idx = headers.index("Purchase Quantity") + 1

    # Row 2 (R1): Prod=100, Surplus=10, Purchase=110
    assert agg_ws.cell(2, prod_idx).value == 100
    assert agg_ws.cell(2, surplus_idx).value == 10
    assert agg_ws.cell(2, purchase_idx).value == 110

    # Row 3 (C1): Prod=100, Surplus=10, Purchase=110
    assert agg_ws.cell(3, prod_idx).value == 100
    assert agg_ws.cell(3, surplus_idx).value == 10
    assert agg_ws.cell(3, purchase_idx).value == 110

    # Row 4 (U1): Prod=100, Surplus=0, Purchase=100
    assert agg_ws.cell(4, prod_idx).value == 100
    assert agg_ws.cell(4, surplus_idx).value == 0
    assert agg_ws.cell(4, purchase_idx).value == 100

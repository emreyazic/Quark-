import openpyxl

from core.bom_parser import BomParser


def test_parse_component_order_list_format_from_real_workbook(tmp_path):
    path = tmp_path / "Component_Order_List_r3.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Components"
    sheet.append([
        "Kart",
        "Manufacturer",
        "Manufacturer Part Number",
        "Quantity",
        "Description",
    ])
    sheet.append(["A", "KEMET", "C0603C104K5RACTU", 4, "100nF capacitor"])
    workbook.save(path)

    parser = BomParser()
    bom = parser.load_file(str(path))
    
    assert bom.column_mapping.is_valid()
    assert bom.headers[bom.column_mapping.mpn] == "Manufacturer Part Number"
    assert bom.headers[bom.column_mapping.manufacturer] == "Manufacturer"
    assert bom.headers[bom.column_mapping.board_identifier] == "Kart"
    
    items = parser.parse_bom_items(bom)
    assert len(items) == 1
    assert items[0].board_name == "Board A"
    assert items[0].mpn == "C0603C104K5RACTU"
    assert items[0].quantity == 4


def test_parse_pcb_kart_format_from_real_workbook(tmp_path):
    path = tmp_path / "XXXX. PCB KART.xlsx"
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Project", "PCB KART"])
    bom_sheet = workbook.create_sheet("BOM")
    bom_sheet.append(["Description", "Qty", "MPN", "Manufacturer"])
    bom_sheet.append(["100nF capacitor", 2, "GRM188R72A104KA35D", "Murata"])
    workbook.save(path)

    parser = BomParser()
    bom = parser.load_file(str(path))

    assert bom.sheet_name == "BOM"
    assert bom.column_mapping.is_valid()
    
    items = parser.parse_bom_items(bom)
    assert len(items) == 1
    assert items[0].mpn == "GRM188R72A104KA35D"
    assert items[0].quantity == 2

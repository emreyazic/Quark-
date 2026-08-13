from openpyxl import Workbook
import threading
import time

from core.database_manager import DatabaseManager
import ui.main_window as main_window_module
from ui.main_window import ComponentLibraryImportWorker


def test_component_library_reader_uses_altium_columns_and_skips_invalid_rows(tmp_path):
    path = tmp_path / "library.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Components"
    sheet.append(["LIBRARYREFERENCE", "MANUFACTURER PART NUMBER", "DIGI-KEY PART NUMBER"])
    sheet.append(["CAP010001", "GRM123", "OLD-DK"])
    sheet.append(["CAP010002", "*", "-"])
    sheet.append(["", "VALID-BUT-NO-CODE", ""])
    sheet.append(["CAP010001", "GRM123", "OLD-DK"])
    workbook.save(path)

    worker = ComponentLibraryImportWorker(str(path), DatabaseManager(str(tmp_path / "db.sqlite")))
    components, skipped = worker._read_components()

    assert components == [("CAP010001", "GRM123")]
    assert skipped == 3


def test_component_library_reader_accepts_comment_as_internal_code(tmp_path):
    path = tmp_path / "library.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["COMMENT", "MANUFACTURER PART NUMBER"])
    sheet.append(["IC010001", "STM32F103"])
    workbook.save(path)

    worker = ComponentLibraryImportWorker(str(path), DatabaseManager(str(tmp_path / "db.sqlite")))
    components, skipped = worker._read_components()

    assert components == [("IC010001", "STM32F103")]
    assert skipped == 0


def test_component_library_searches_in_parallel_and_reuses_cache(tmp_path, monkeypatch):
    path = tmp_path / "library.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Components"
    sheet.append(["LIBRARYREFERENCE", "MANUFACTURER PART NUMBER"])
    for index in range(12):
        sheet.append([f"IC{index}", f"MPN{index}"])
    workbook.save(path)

    lock = threading.Lock()
    calls = {"lcsc": 0, "digikey": 0}

    class FakeJlcSearcher:
        def __init__(self, *args, **kwargs):
            pass

        def _resolve_lcsc_from_mpn(self, mpn):
            with lock:
                calls["lcsc"] += 1
            time.sleep(0.03)
            return f"C{mpn[3:]}"

        def close(self):
            pass

    class FakeDigiKeySearcher:
        def __init__(self, *args, **kwargs):
            pass

        def search_mpn(self, mpn, required_stock=1, include_live_data=True):
            with lock:
                calls["digikey"] += 1
            time.sleep(0.03)
            return type("Result", (), {
                "configured": True,
                "error": None,
                "digikey_part_number": f"DK-{mpn}",
            })()

        def close(self):
            pass

    monkeypatch.setattr(main_window_module, "JlcpcbSearcher", FakeJlcSearcher)
    monkeypatch.setattr(main_window_module, "DigiKeySearcher", FakeDigiKeySearcher)

    db = DatabaseManager(str(tmp_path / "db.sqlite"))
    first = ComponentLibraryImportWorker(str(path), db)
    started = time.monotonic()
    first.run()
    elapsed = time.monotonic() - started

    assert calls == {"lcsc": 12, "digikey": 12}
    assert elapsed < 0.60

    second = ComponentLibraryImportWorker(str(path), db)
    second.run()
    assert calls == {"lcsc": 12, "digikey": 12}

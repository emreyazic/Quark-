import openpyxl
import pytest

from core.bom_parser import BomParser
from models.bom_item import BomFile, ColumnMapping


def test_parser_parses_only_targeted_sheet_without_silent_combination(tmp_path):
    """Verify that parse_bom_items parses only the targeted sheet and does not combine rows from other sheets."""
    path = tmp_path / "multi-sheet.xlsx"
    workbook = openpyxl.Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Project", "Example"])

    board_a = workbook.create_sheet("Board A")
    board_a.append(["MPN", "Quantity", "Description"])
    board_a.append(["MPN-A", 2, "Part A"])

    board_b = workbook.create_sheet("Board B")
    board_b.append(["Description", "Qty", "Manufacturer Part Number"])
    board_b.append(["Part B", 3, "MPN-B"])

    hidden_helper = workbook.create_sheet("Helper")
    hidden_helper.append(["MPN", "Quantity"])
    hidden_helper.append(["SHOULD-NOT-LOAD", 1])
    hidden_helper.sheet_state = "hidden"
    workbook.save(path)

    parser = BomParser()

    # Load Board A
    bom_file_a = parser.load_file(str(path), sheet_name="Board A")
    items_a = parser.parse_bom_items(bom_file_a)

    assert bom_file_a.sheet_name == "Board A"
    # Must only contain items from Board A, NOT Board B!
    assert [(item.mpn, item.quantity) for item in items_a] == [
        ("MPN-A", 2),
    ]

    # Load Board B
    bom_file_b = parser.load_file(str(path), sheet_name="Board B")
    items_b = parser.parse_bom_items(bom_file_b)

    assert bom_file_b.sheet_name == "Board B"
    assert [(item.mpn, item.quantity) for item in items_b] == [
        ("MPN-B", 3),
    ]


def test_inspect_sheets_detects_all_visible_sheets(tmp_path):
    """inspect_sheets should return metadata for all visible worksheets."""
    path = tmp_path / "inspect.xlsx"
    workbook = openpyxl.Workbook()

    ws1 = workbook.active
    ws1.title = "Power Supply"
    ws1.append(["MPN", "Qty", "Designator"])
    ws1.append(["LM317T", 1, "U1"])
    ws1.append(["100uF", 2, "C1, C2"])

    ws2 = workbook.create_sheet("MCU Board")
    ws2.append(["Manufacturer Part Number", "Quantity"])
    ws2.append(["STM32F401", 1])

    ws_hidden = workbook.create_sheet("HiddenNotes")
    ws_hidden.append(["Notes", "Data"])
    ws_hidden.sheet_state = "hidden"

    workbook.save(path)

    parser = BomParser()
    sheets = parser.inspect_sheets(str(path))

    assert len(sheets) == 2
    sheet_names = [s.sheet_name for s in sheets]
    assert "Power Supply" in sheet_names
    assert "MCU Board" in sheet_names
    assert "HiddenNotes" not in sheet_names

    ps_sheet = next(s for s in sheets if s.sheet_name == "Power Supply")
    assert ps_sheet.row_count == 2
    assert ps_sheet.column_mapping.mpn == 0
    assert ps_sheet.column_mapping.quantity == 1
    assert ps_sheet.column_mapping.designator == 2
    assert ps_sheet.is_valid is True


def test_inspect_sheets_detects_duplicate_sheets(tmp_path):
    """inspect_sheets should flag exact duplicate and copy sheets with warnings."""
    path = tmp_path / "duplicates.xlsx"
    workbook = openpyxl.Workbook()

    ws1 = workbook.active
    ws1.title = "Board A"
    ws1.append(["MPN", "Quantity", "Description"])
    ws1.append(["RES-10K", 10, "Resistor 10k"])
    ws1.append(["CAP-100N", 5, "Capacitor 100nF"])

    ws2 = workbook.create_sheet("Board A (Copy)")
    ws2.append(["MPN", "Quantity", "Description"])
    ws2.append(["RES-10K", 10, "Resistor 10k"])
    ws2.append(["CAP-100N", 5, "Capacitor 100nF"])

    ws3 = workbook.create_sheet("Board B")
    ws3.append(["MPN", "Quantity"])
    ws3.append(["MCU-01", 1])

    workbook.save(path)

    parser = BomParser()
    sheets = parser.inspect_sheets(str(path))

    assert len(sheets) == 3
    sheet_a = next(s for s in sheets if s.sheet_name == "Board A")
    sheet_copy = next(s for s in sheets if s.sheet_name == "Board A (Copy)")
    sheet_b = next(s for s in sheets if s.sheet_name == "Board B")

    assert sheet_a.duplicate_of is None
    assert sheet_copy.duplicate_of == "Board A"
    assert any("Probable duplicate" in w for w in sheet_copy.warnings)
    assert sheet_b.duplicate_of is None


def test_parser_rejects_legacy_xls_with_clear_message(tmp_path):
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"not-an-xls-workbook")

    with pytest.raises(ValueError, match=r"Legacy \.xls files are not supported"):
        BomParser().load_file(str(path))

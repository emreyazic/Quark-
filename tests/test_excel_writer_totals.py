import openpyxl

from core.excel_writer import ExcelWriter
from models.bom_item import BomItem


def test_unit_mode_total_cost_sums_extended_price_columns(tmp_path):
    items = [
        BomItem(
            mpn="A",
            pricing_quantity=10,
            unit_price=2.0,
            digikey_unit_price=3.0,
            jlcpcb_total_price=20.0,
            digikey_total_price=30.0,
        ),
        BomItem(
            mpn="B",
            pricing_quantity=10,
            unit_price=4.0,
            digikey_unit_price=5.0,
            jlcpcb_total_price=40.0,
            digikey_total_price=50.0,
        ),
    ]
    output_path = tmp_path / "unit-totals.xlsx"

    ExcelWriter(items, pricing_mode="unit", build_multipliers=[1]).write(
        str(output_path)
    )

    workbook = openpyxl.load_workbook(output_path, data_only=False)
    sheet = workbook["Enriched BOM"]
    headers = [cell.value for cell in sheet[1]]
    total_row = len(items) + 2
    label_col = headers.index("Pricing Quantity") + 1
    jlc_total_col = headers.index("JLCPCB Total Price") + 1
    dk_total_col = headers.index("DigiKey Total Price") + 1
    jlc_unit_col = headers.index("JLCPCB Unit Price") + 1

    assert sheet.cell(total_row, label_col).value == "Total Cost"
    assert sheet.cell(total_row, jlc_total_col).value == "=SUM(P2:P3)"
    assert sheet.cell(total_row, dk_total_col).value == "=SUM(Q2:Q3)"
    assert sheet.cell(total_row, jlc_unit_col).value is None


def test_cost_sheet_uses_scalar_prices_when_tiers_are_empty_or_invalid():
    item = BomItem(
        mpn="A",
        quantity=2,
        jlcpcb_part_number="C1",
        unit_price=1.5,
        digikey_unit_price=2.5,
        jlcpcb_price_breaks_raw="not-json",
        digikey_price_breaks=[],
    )
    writer = ExcelWriter([item], pricing_mode="project", build_multipliers=[1])

    writer._write_single_cost_sheet("Cost", [item])

    sheet = writer.wb["Cost"]
    assert sheet["B2"].value == 3.0
    assert sheet["E2"].value == 5.0
    assert sheet["H5"].value == 1.5
    assert sheet["I5"].value == 2.5

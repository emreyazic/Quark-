import os
import openpyxl
import pytest

from models.bom_item import BomItem
from models.project import Project, ProjectItem
from services.project_aggregation import aggregate_project
from core.project_excel_writer import ProjectExcelWriter

def test_workbook_structure(tmp_path):
    project = Project("Test")
    project.add_board(ProjectItem("path_a", "Board A", 2))
    
    # Just one empty board
    agg = aggregate_project(project)
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=[],
        component_keys=[]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    sheet_names = wb.sheetnames
    
    assert "Master Summary" in sheet_names
    assert "Aggregated Components" in sheet_names
    assert "Board A" in sheet_names
    assert "Raw - Board A" in sheet_names

def test_master_summary_totals_and_unpriced(tmp_path):
    project = Project("Test")
    board = ProjectItem("path_a", "Board A", 2)
    board.bom_items = [
        BomItem(mpn="R1", quantity=5, jlcpcb_part_number="C123"), # JLC
        BomItem(mpn="R2", quantity=3), # DigiKey only
        BomItem(mpn="R3", quantity=1), # Unpriced
    ]
    project.add_board(board)
    agg = aggregate_project(project)
    
    # Enrich them
    enriched = [
        BomItem(mpn="R1", jlcpcb_part_number="C123", status="✅ Found", jlcpcb_price_breaks_raw='[{"qFrom": 1, "price": 0.10}]'),
        BomItem(mpn="R2", status="Not found", digikey_price_breaks=[(1, 0.50)]),
        BomItem(mpn="R3", status="Not found") # Unpriced
    ]
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=enriched,
        component_keys=[c.component_key for c in agg.components],
        build_multipliers=[1]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Master Summary"]
    
    # R1: 5 * 2 = 10 qty * 0.10 = $1.00 JLCPCB
    # R2: 3 * 2 = 6 qty * 0.50 = $3.00 DigiKey
    # R3: 1 * 2 = 2 qty * no price = $0
    
    found_totals = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "1x":
            jlc_cost = row[1]
            dk_cost = row[2]
            comb = row[3]
            dk_only = row[4]
            
            assert jlc_cost == 1.0
            assert dk_cost == 3.0
            assert comb == 4.0
            assert dk_only == 3.0
            found_totals = True
            
    assert found_totals

def test_unpriced_in_aggregated_components_sheet(tmp_path):
    project = Project("Test")
    board = ProjectItem("path_a", "Board A", 2)
    board.bom_items = [
        BomItem(mpn="R3", quantity=1), # Unpriced
    ]
    project.add_board(board)
    agg = aggregate_project(project)
    
    enriched = [
        BomItem(mpn="R3", status="Not found") # Unpriced
    ]
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=enriched,
        component_keys=[c.component_key for c in agg.components],
        build_multipliers=[1]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Aggregated Components"]
    
    headers = [cell.value for cell in ws[1]]
    
    # Regression test for absent bloated headers
    assert "Selected Source @ 5x" not in headers
    assert "Selected Source @ 1x" not in headers
    assert "JLCPCB Extended Cost @ 5x" not in headers
    
    mpn_col_idx = headers.index("MPN")
    jlcpcb_price_col_idx = headers.index("JLCPCB Unit Price")
    digikey_price_col_idx = headers.index("DigiKey Unit Price")
    status_col_idx = headers.index("Status")
    
    found_unpriced = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[mpn_col_idx] == "R3":
            assert row[jlcpcb_price_col_idx] is None
            assert row[digikey_price_col_idx] is None
            assert row[status_col_idx] == "Not found"
            found_unpriced = True
            
    assert found_unpriced

def test_jlcpcb_negative_status_fallback_to_digikey(tmp_path):
    project = Project("Test")
    board = ProjectItem("path_a", "Board A", 1)
    board.bom_items = [
        BomItem(mpn="R1", quantity=1),
    ]
    project.add_board(board)
    agg = aggregate_project(project)
    
    enriched = [
        BomItem(
            mpn="R1", 
            jlcpcb_part_number="C123", 
            status="Insufficient JLCPCB stock", 
            jlcpcb_price_breaks_raw='[{"qFrom": 1, "price": 0.10}]',
            digikey_price_breaks=[(1, 0.50)]
        )
    ]
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=enriched,
        component_keys=[c.component_key for c in agg.components],
        build_multipliers=[1]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Master Summary"]
    
    for row in ws.iter_rows(values_only=True):
        if row[0] == "1x":
            jlc_cost = row[1]
            dk_cost = row[2]
            comb = row[3]
            dk_only = row[4]
            
            # Since JLCPCB has a negative status, it should fallback to DK
            assert jlc_cost == 0.0
            assert dk_cost == 0.50
            assert comb == 0.50
            assert dk_only == 0.50

def test_board_quantity_math(tmp_path):
    project = Project("Test")
    board = ProjectItem("path_a", "Board A", 5) # 5 boards
    board.bom_items = [
        BomItem(mpn="R1", quantity=3) # 3 per board
    ]
    project.add_board(board)
    agg = aggregate_project(project)
    
    enriched = [
        BomItem(mpn="R1", digikey_price_breaks=[(1, 0.20)])
    ]
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=enriched,
        component_keys=[c.component_key for c in agg.components],
        build_multipliers=[1, 10]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    board_ws = wb["Board A"]
    
    headers = [cell.value for cell in board_ws[6]]
    qty_col = headers.index("Quantity (Per Board / Total)")
    data_row = [cell.value for cell in board_ws[7]]
    assert data_row[qty_col] == "3 / 15"
    
    # 1 product = 5 boards * 3 qty * $0.20 = 15 * $0.20 = $3.00
    # 10 products = 10 * $3.00 = $30.00
    summary_ws = wb["Master Summary"]
    totals = {row[0]: row for row in summary_ws.iter_rows(values_only=True) if row[0] in ("1x", "10x")}
    assert totals["1x"][2] == 3.0
    assert totals["1x"][3] == 3.0
    assert totals["10x"][2] == 30.0
    assert totals["10x"][3] == 30.0

def test_multiboard_quantity_formatting_and_auto_width(tmp_path):
    long_desc = "Precision resistor network for controller power measurement"
    project = Project("Test")
    
    board_a = ProjectItem("path_a", "Board A", 2)
    board_a.bom_items = [
        BomItem(mpn="R1", quantity=2.0, description=long_desc, designator="R1")
    ]
    board_b = ProjectItem("path_b", "Board B", 1)
    board_b.bom_items = [
        BomItem(mpn="R1", quantity="3.0", description=long_desc, designator="R2")
    ]
    project.add_board(board_a)
    project.add_board(board_b)
    agg = aggregate_project(project)
    
    enriched = [
        BomItem(mpn="R1", description=long_desc)
    ]
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=enriched,
        component_keys=[c.component_key for c in agg.components],
        build_multipliers=[1]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    agg_ws = wb["Aggregated Components"]
    
    assert agg_ws["A2"].value == "Board A: 4\nBoard B: 3"
    assert agg_ws["D2"].value == "2 / 4\n3 / 3"
    assert agg_ws.column_dimensions["B"].width >= len(long_desc)
    
    board_ws = wb["Board A"]
    assert board_ws["D7"].value == "2 / 4"
    
    raw_ws = wb["Raw - Board A"]
    assert raw_ws["D2"].value == "2"

def test_warnings_included(tmp_path):
    project = Project("Test")
    board = ProjectItem("path_a", "Board A", 2)
    board.bom_items = [
        BomItem(mpn="R1", quantity="abc") # Invalid
    ]
    project.add_board(board)
    agg = aggregate_project(project)
    
    writer = ProjectExcelWriter(
        project=project,
        aggregation_result=agg,
        enriched_items=[],
        component_keys=[]
    )
    
    out_path = tmp_path / "out.xlsx"
    writer.write(str(out_path))
    
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Master Summary"]
    
    found_warning = False
    for row in ws.iter_rows(values_only=True):
        if row[0] and "Skipped" in str(row[0]) and "abc" in str(row[0]):
            found_warning = True
            
    assert found_warning
